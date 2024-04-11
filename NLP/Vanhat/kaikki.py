import asyncio
import threading
import hae
import nltk
import openpyxl as xl
import os
import pandas as pd
import queue
import spacy
import statistics
import time
import tokenisointi
import tkinter as tk
import ttkbootstrap as ttk
import uralic

from collections import Counter
from functools import partial
from itertools import groupby
from nltk import Counter
from openpyxl.utils.dataframe import dataframe_to_rows
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from threading import Thread
from tkinter import Menu, ttk as tkstyle
from tokenisointi import tokenisoi_sanat
from ttkbootstrap.dialogs import Messagebox
from ttkbootstrap.constants import *
from uralicNLP import uralicApi

# ----------- GLOBAALIT MUUTTUJAT JA FUNKTIOT-----------
eka_taso = "  "
toka_taso = "    "
kolmas_taso = "      "

# Asetussanakirjan alustus
asetukset = {}

korpus = {}

# Haetaan korpus ja lasketaan sen koko
def alusta():
    global korpus
    global prog_yht
     # Noudetaan korpus
    polku = os.path.abspath(os.path.curdir)
    kansio = "corpora"
    korpus = hae.hae_korpus(os.path.join(polku, kansio))
    prog_yht = hae_kasiteltavat_yht(korpus)

# Asetusten päivittämisfunktio
def valitse(arvo, asetus):
    global asetukset
    asetukset[asetus] = arvo.get()

# Funktio, jossa lasketaan kaikkien käsiteltävien entiteettien määrä,
# eli kansiot ja niissä olevat tiedostot
def hae_kasiteltavat_yht(dct, yht = 0):
    yht = yht + len(dct)
    if isinstance(dct, dict):
        for avain1, arvo1 in dct.items():
            if isinstance(arvo1, dict):
                yht = hae_kasiteltavat_yht(arvo1, yht)
    print("Tiedostoja: ", yht)
    return yht

# Funktio, jossa lasketaan asetusten perusteella, montako laskua/vaihetta
# ohjelmassa on. Esimerkiksi käyttäjä on valinnut kohdat Sanamäärä,
# Virkemäärä ja kosinisimilaarisuus jolloin laskuja on hae_kasiteltavat_yht() x 2 + 
# len(korpus), koska viimeisessä laskuissa käsitellään koko kirjoittajaa,
# ei yksittäisiä tekstejä. 
def hae_laskut_yht():
    global korpus
    yht = 0
    kasiteltavia = hae_kasiteltavat_yht(korpus)
    for asetus in asetukset:
        if asetukset[asetus] == True:
            if asetus == "Kosinisimilaarisuus" or asetus == "HFL 50" or asetus == "HFL 100":
                yht += len(korpus)
            else:
                yht += kasiteltavia
    print("Vaiheita: ", yht)
    return yht

# ----------- GUI-----------

class Splash(ttk.Toplevel):
    def __init__(self, parent):
        ttk.Toplevel.__init__(self, parent)
        self.title("Splash")

        self.position_center()
        self.geometry("200x200")

        teksti = "Tervetuloa! Pieni hetki, sovellus käynnistyy."
        splash_label = ttk.Text(self, wrap=WORD)
        splash_label.insert(INSERT, teksti)
        splash_label.pack()

        self.update()

class Vasen:
    # Edistymispalkin muuttujat
    prog_yht = 0
    prog_valmiita = 0
    
    def __init__(self, parent, vasen_grid):
        self.parent = parent
        self.vasen_grid = vasen_grid

        ttk.Label(vasen_grid, text="Asetukset", font=("Garamond", 28)).grid(
        row=0, columnspan=2
    )
        # ----------- PERUSTILASTOT -----------
        perustilastot = [
            "Sanamäärä",
            "Virkemäärä",
            "Grafeemeja sanoissa ja virkkeissä",
            "Morfeemeja sanoissa",
            "Virkepituus sanoina",
        ]

        # Luodaan Frame Framen sisään osiolle
        tilastot_frame = ttk.Frame(vasen_grid)
        tilastot_frame.grid(row=1, column=1, pady=10, padx=10, sticky="w")

        perustilastot_label = ttk.Label(
            vasen_grid, text="Perustilastoja", font=("Garamond 15 italic")
        )
        perustilastot_label.grid(row=1, column=0)

        self.luo_valintalaatikot(tilastot_frame, perustilastot)

        # ----------- SANASTOASETUKSET -----------
        sanasto = [
            "Type-token ratio",
            "Sanaston tiheys",
            "Sanaluokkien frekvenssit",
            "HFL 50",
            "HFL 100",
            "Hapax Legomena",
            "Erisnimet",
        ]

        # Luodaan Frame Framen sisään osiolle
        sanasto_frame = ttk.Frame(vasen_grid)
        sanasto_frame.grid(row=2, column=1, pady=10, padx=10, sticky="w")

        sanasto_label = ttk.Label(
            vasen_grid, text="Sanastotilastoja", font=("Garamond 15 italic")
        )
        sanasto_label.grid(row=2, column=0)

        self.luo_valintalaatikot(sanasto_frame, sanasto)

        # ----------- KIRJOITTAJIEN VERTAILU -----------
        vertailu = ["Yule K", "Yule I", "Kosinisimilaarisuus"]

        # Luodaan Frame Framen sisään osiolle
        vertailu_frame = ttk.Frame(vasen_grid)
        vertailu_frame.grid(row=3, column=1, pady=10, padx=10, sticky="w")

        vertailu_label = ttk.Label(
            vasen_grid, text="Kirjoittajien vertailu", font=("Garamond 15 italic")
        )
        vertailu_label.grid(row=3, column=0)

        self.luo_valintalaatikot(vertailu_frame, vertailu)

        # ----------- KÄYNNISTÄ -----------

        kaynnista_frame = ttk.Frame(vasen_grid, bootstyle="primary")
        kaynnista_frame.grid(row=4, column=0, pady=20, columnspan=2)

        self.kaynnista_nappi = ttk.Button(
            kaynnista_frame,
            text="Käynnistä",
            command=self.kaynnista,
        )
        self.kaynnista_nappi.pack(anchor="w")

        # ----------- EDISTYMISPALKKI -----------
        ttk.Style().configure("custom.Horizontal.TProgressbar", 
                              thickness=15, 
                              bordercolor="black",
                              borderwidth="1",
                              background="brown",
                              troughcolor="beige",
                              pbarrelief='solid')      
        edistymis_frame = ttk.Frame(vasen_grid)
        edistymis_frame.grid(row=5, column=0, pady=20, columnspan=2)
        self.edistymispalkki = ttk.Progressbar(edistymis_frame,
                                        mode="determinate", 
                                        orient="horizontal",
                                        length=200,
                                        style="custom.Horizontal.TProgressbar")

        self.edistymispalkki.pack_forget()        
        
        # ----------- FUNKTIOT -----------

        # Kaikkien osioiden valintalaatikot luodaan
        # listasta silmukassa.
    def luo_valintalaatikot(self, sijainti, vaihtoehdot):
        global asetukset
        for vaihtoehto in vaihtoehdot:
            asetukset[vaihtoehto] = False  # Alustetaan kaikki sanakirjaan
            checkbox_var = tk.BooleanVar()
            checkbox_var.set(False)
            ttk.Checkbutton(
                sijainti,
                variable=checkbox_var,
                text=vaihtoehto,
                command=partial(valitse, checkbox_var, vaihtoehto),
            ).pack(anchor="w")
    
    def kaynnista(self):
        print("Käynnistä")
        self.nayta_edistymispalkki
        saie = Thread(target=self.aloita_ohjelma, 
                           daemon=True)
        saie.start()
                
    def aloita_ohjelma(self):
        print("Aloita ohjelma")
            # Ohjelman suorituksen ajaksi otetaan käynnistysnappi pois käytöstä
        self.kaynnista_nappi.config(state="disabled")
        self.edistymispalkki.config(value=Vasen.prog_valmiita)
        Openpyxl.hae_workbook()
        Openpyxl.kaynnista_ohjelma(asetukset)
        self.kaynnista_nappi.config(state="normal")

    def paivita_palkki():
        Vasen.prog_valmiita += 1
        Vasen.edistymispalkki.config(value=Vasen.prog_valmiita)

    def nayta_edistymispalkki(self):
        vaiheita = hae_laskut_yht()
        self.edistymispalkki.config(maximum=vaiheita, value=10)
        self.edistymispalkki.pack()

class Oikea:
    def __init__(self, parent, oikea_grid):
        self.parent = parent
        self.oikea_grid = oikea_grid
        ttk.Label(oikea_grid, text="Ohjeet", font=("Garamond", 28)).grid(row=0)

        ohje = """Helmin NLP-ohjelma on työkalu suomenkielisten tekstien yksinkertaiseen analyysiin ja vertailuun. 
                      
Ohjelman suorituskansiosta löytyy "corpora"-niminen kansio, johon voit tiputtaa haluamasi UTF-8 -muotoiset tekstit. Mikäli haluat vertailla useampaa kirjoittajaa, luo jokaiselle kirjailijalle oma alikansio. Älä kuitenkaan luo kirjailijan alakansion sisälle sisempiä alakansioita. Kansioiden ja tiedostojen nimillä ei ole merkitystä. 
                      
Kaikki tilastot ja vertailujen tulokset tallennetaan samaan, myöskin ohjelman suorituskansiosta löytyvään tiedostoon nimeltä "tulokset.xlsx."

Tulevia toimintoja:
    - lauseiden tunnistaminen ja analyysi
    - latauspalkki
    - painike ohjelman suorituksen keskeyttämiseksi               
                      """

        teksti = ttk.Text(
            oikea_grid,
            wrap=WORD,
        )
        teksti.grid(row=1)
        teksti.insert(END, ohje)
        teksti.config(state=DISABLED)  # Vain-luku -tilainen teksti

class Sovellus(ttk.Window):

    def __init__(self):
        super().__init__(themename="minty")
        self.withdraw()
        self.splash = Splash(self)

        # Alustetaan ikkuna
        self.title("Helmin NLP-ohjelma")
        self.iconbitmap("99_85283.ico")

        # Alustetaan ikkunan koko suhteessa näytön kokoon
        leveys = self.winfo_screenwidth()
        korkeus = self.winfo_screenheight()
        self.position_center()
        self.geometry("%sx%s" % (int(leveys / 1.5), int(korkeus / 1.5)))

        # Haetaan korpus
        alusta()

        #time.sleep(4)
        # Odotetaan hetki, jotta kaikki varmasti on valmista
        self.splash.destroy()
        self.deiconify()

        # ------------ PÄÄIKKUNAN FUNKTIOT ------------

        # Funktio sovelluksen teeman vaihtamiseen
        self.tyyli = ttk.Style()

        def vaihdavari(vari):
            self.tyyli.theme_use(vari)

        def naytainfo():
            self.Messagebox.ok(
                """Sovelluksessa käytetyt kirjastot:
NLTK, Openypxl, Pandas, Scikit-learn, Spacy, TTKBootstrap, UralicNLP.
Github: https://github.com/ellaelomaa/nlp
Tekijät: Ella Elomaa ja Helmi Siiroinen (2024)
"""
            )

        # ----------- TYÖKALUPALKKI ------------

        self.menubar = Menu(self)
        self.config(menu=self.menubar)

        # Infoikkuna

        self.info_vaihtoehdot = Menu(self.menubar)
        self.info_vaihtoehdot.add_command(label="Käyttöohje")
        self.info_vaihtoehdot.add_command(label="Tietoa sovelluksesta", command=naytainfo)

        # Työkalupalkin sisältö
        self.menubar.add_cascade(label="Info", menu=self.info_vaihtoehdot)

        # Sovelluksen teeman vaihto
        self.varit = Menu(self.menubar)
        for x in ["flatly", "minty", "morph", "solar", "superhero", "darkly"]:
            self.varit.add_command(label=x, command=lambda x=x: vaihdavari(x))

        self.menubar.add_cascade(label="Teemat", menu=self.varit)

        # ----------- KÄYTTÖÖLIITTYMÄN PÄÄRAKENNE ----------
        # Ikkuna jakautuu kahteen puoliskoon: vasemmalla on asetukset,
        # oikealla tekstien käsittelyn tuloslaatikko

        self.vasen_frame = ttk.Frame(self)
        self.vasen_frame.grid(row=0, column=0, sticky="ns", pady=50)

        self.vasen_sisalto = Vasen(self, self.vasen_frame)

        self.oikea_frame = ttk.Frame(self)
        self.oikea_frame.grid(row=0, column=1, sticky="ns", pady=50)

        self.oikea_sisalto = Oikea(self, self.oikea_frame)

        # Molemmat puoliskojen paino on sama, eli molemmat saavat 50 % ikkunan leveydestä
        self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=1)


class Openpyxl:
    sarakesanakirja = {}
    rivisanakirja = {}
    # ----------- OPENPYXL-----------

    def hae_workbook():

        # Tarkistetaan, onko tiedosto jo olemassa.
        polku = "./tulokset.xlsx"

        if os.path.isfile(polku) == True:
            workbook = xl.load_workbook("tulokset.xlsx")
            return workbook
        else: # Jos ei, niin luodaan tiedosto ja sen sivut
            workbook = xl.Workbook()

            # Korvataan ensimmäisen sivun oletusnimi
            teksteittain = workbook["Sheet"]
            teksteittain.title = "Teksteittäin"

            # Muut sivut
            workbook.create_sheet("HFL")
            workbook.create_sheet("Kosinisimilaarisuus")
            return workbook

    # Sanakirjat rivien ja sarakkeiden nimille ja niiden indekseille.
    # OpenPyxl:ssä tiettyyn soluun voi lisätä rivin ja sarakkeen indeksin
    # mukaan, ei nimen, joten luodaan sanakirja muotoa nimi:indeksi sekä
    # riveille että sarakkeille.
    # rivisanakirja = {}
    # sarakesanakirja = {}

    # Tulostiedoston rivien otsikot
    rivinimet = [
        "Sanamäärä",
        "Virkemäärä",
        "Grafeemeja sanoissa",
        "Grafeemeja virkkeissä",
        "Morfeemeja (avg)",
        "Morfeemeja (mediaani)",
        "Morfeemeja (moodi)",
        "Morfeemeja (min)",
        "Morfeemeja (max)",
        "Morfeemeja (vaihteluväli)",
        "Morfeemeja (keskihajonta)",
        "Morfeemeja (varianssi)",
        "Virkepituus sanoina (avg)",
        "Virkepituus sanoina (keskihajonta)",
        "Saneita",
        "Tyyppejä",
        "TTR",
        "Sanaston tiheys",
        "Funktiosanoja",
        "Funktiosanoja (%)",
        "Hapax legomenon",
        "Yule I",
        "Yule K",
    ]

    # ----------- WORKBOOK-FUNKTIOT -----------

    def hae_tekstinimet(korpus):
        nimet = []
        for kirjoittaja in korpus:
            for teksti in korpus[kirjoittaja]:
                nimet.append(teksti)
            nimet.append(kirjoittaja)
        return nimet

    def luo_sarakenimet(korpus):
        sarakkeet = {}
        nykyinen = 2
        for kirjoittaja in korpus:
            for teksti in korpus[kirjoittaja]:
                sarakkeet[teksti] = nykyinen
                nykyinen += 1
            sarakkeet[kirjoittaja] = nykyinen
            nykyinen += 1

        return sarakkeet

    def luo_rivinimet():
        rivit = {}
        nykyinen = 2

        for rivinimi in Openpyxl.rivinimet:
            rivit[rivinimi] = nykyinen
            nykyinen += 1
        return rivit

    # Luodaan taulukon perusmuoto
    def alusta_taulukko(korpus):
        workbook = Openpyxl.hae_workbook()
        workbook.active = workbook["Teksteittäin"]
        sivu = workbook.active

        # Alustetaan rivien (eli asetusten) nimet
        rivinro = 2
        for rivi in Openpyxl.rivinimet:
            solu = sivu.cell(row=rivinro, column=1)
            solu.value = rivi
            rivinro += 1

        # Alustetaan sarakkeiden (eli tekstien ja kirjoittajien) nimet
        otsikot = Openpyxl.hae_tekstinimet(korpus)
        sarakenro = 2
        for sarake in otsikot:
            solu = sivu.cell(row=1, column=sarakenro)
            solu.value = sarake
            sarakenro += 1

    def alusta_hfl_sivu(korpus):
        workbook = Openpyxl.hae_workbook()
        workbook.active = workbook["HFL"]
        sivu = workbook.active
        # Alustetaan sarakkeiden (eli tekstien ja kirjoittajien) nimet
        otsikot = Openpyxl.hae_tekstinimet(korpus)
        sarakenro = 1
        for sarake in otsikot:
            solu = sivu.cell(row=1, column=sarakenro)
            solu.value = sarake
            sarakenro += 1

    def lisaa_taulukkoon(sivunimi, arvo, rivi, sarake):
        workbook = Openpyxl.hae_workbook()
        workbook.active = workbook[sivunimi]
        sivu = workbook.active
        solu = sivu.cell(
            row=Openpyxl.rivisanakirja[rivi], column=Openpyxl.sarakesanakirja[sarake]
        )
        solu.value = arvo

    def lisaa_hfl_sivulle(arvo, rivi, sarake):
        workbook = Openpyxl.hae_workbook()
        workbook.active = workbook["HFL"]
        sivu = workbook.active
        solu = sivu.cell(row=rivi, column=Openpyxl.sarakesanakirja[sarake])

        solu.value = arvo

    def df_taulukkoon(sivunimi, data):
        workbook = Openpyxl.hae_workbook()
        workbook.active = workbook[sivunimi]
        sivu = workbook.active
        for rivi in dataframe_to_rows(data, index=True, header=True):
            sivu.append(rivi)

    def kaynnista_ohjelma(asetukset):
        workbook = Openpyxl.hae_workbook()

        # Alustetaan sarakkeiden/rivien nimet sanakirjaan, jotta tietoja voidaan
        # tallentaa tiettyyn sarakkeeseen (tekstiin/kirjoittajaan) tai riviin.
        # Solun arvon saa komennolla row_cells[sarakenimet["Jokin"]].value
        Openpyxl.sarakesanakirja = Openpyxl.luo_sarakenimet(korpus)
        Openpyxl.rivisanakirja = Openpyxl.luo_rivinimet()
        Openpyxl.alusta_taulukko(korpus)
        Openpyxl.alusta_hfl_sivu(korpus)

        # Käynnistetään ohjelma ja aloitetaan laskut.
        Laskut.tilastoja(asetukset, korpus)

        workbook.save("tulokset.xlsx")
        print("Valmista")


class Laskut:

    # Kokonaissanamäärä per teksti ja kirjoittaja
    def sanamaara(korpus):
        for kirjoittaja in korpus:
            sanojaKirjoittaja = 0
            for teksti in korpus[kirjoittaja]:
                sanojaTekstissa = len(
                    tokenisointi.tokenisoi_sanat(korpus[kirjoittaja][teksti])
                )
                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin", sanojaTekstissa, "Sanamäärä", teksti
                )
                sanojaKirjoittaja += sanojaTekstissa
                Vasen.paivita_palkki()
            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin", sanojaKirjoittaja, "Sanamäärä", kirjoittaja)
            Vasen.paivita_palkki()
    # Virkkeiden määrä per teksti ja kirjoittaja
    def virkemaara(korpus):
        for kirjoittaja in korpus:
            lauseita_kirjoittaja = 0
            for teksti in korpus[kirjoittaja]:
                lauseitaTekstissa = len(
                    tokenisointi.tokenisoi_virkkeet(korpus[kirjoittaja][teksti])
                )
                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin", lauseitaTekstissa, "Virkemäärä", teksti
                )
                lauseita_kirjoittaja += lauseitaTekstissa
            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin", lauseita_kirjoittaja, "Virkemäärä", kirjoittaja
            )

    # Sanojen määrien keskiarvo ja keskihajonta per teksti ja kirjoittaja
    def laske_sanat_virkkeissa(korpus):
        for kirjoittaja in korpus:
            sanoja_kirjoittaja = []
            for teksti in korpus[kirjoittaja]:
                virkkeet = tokenisointi.tokenisoi_virkkeet(korpus[kirjoittaja][teksti])
                sanoja_virkkeissa = []
                for virke in virkkeet:
                    virke = tokenisointi.tokenisoi_sanat(virke)
                    sanoja_virkkeissa.append(len(virke))
                    sanoja_kirjoittaja.append(len(virke))

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    round(statistics.mean(sanoja_virkkeissa), 2),
                    "Virkepituus sanoina (avg)",
                    teksti,
                )
                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    round(statistics.stdev(sanoja_virkkeissa), 2),
                    "Virkepituus sanoina (keskihajonta)",
                    teksti,
                )
            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                round(statistics.mean(sanoja_kirjoittaja), 2),
                "Virkepituus sanoina (avg)",
                kirjoittaja,
            )
            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                round(statistics.stdev(sanoja_kirjoittaja), 2),
                "Virkepituus sanoina (keskihajonta)",
                kirjoittaja,
            )

    def laske_grafeemeina(korpus):
        # Käydään korpus kirjoittaja kerrallaan läpi

        for kirjoittaja in korpus:
            # Alustetaan muuttujat kirjoittajaa koskeville muuttujille.
            sanoja_kirjoittaja = 0
            merkkeja_kirjoittaja = 0

            # Käydään kirjoittajan tekstit kerrallaan läpi
            for teksti in korpus[kirjoittaja]:
                merkkeja_tekstissa = 0

                # Keskiarvoa varten lasketaan sanojen määrä tekstissä
                sanalista = tokenisointi.tokenisoi_sanat(korpus[kirjoittaja][teksti])
                sanoja_tekstissa = len(sanalista)

                # Kasvatetaan samalla koko kirjoittajan sanamäärää laskevaa muuttujaa
                sanoja_kirjoittaja += len(sanalista)

                # Lasketaan, kuinka pitkiä sanat ovat teksteittäin
                for sana in sanalista:
                    merkkeja_tekstissa += len(sana)
                    merkkeja_kirjoittaja += len(sana)

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    round(merkkeja_tekstissa / sanoja_tekstissa, 2),
                    "Grafeemeja sanoissa",
                    teksti,
                )
            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                round(merkkeja_kirjoittaja / sanoja_kirjoittaja, 2),
                "Grafeemeja sanoissa",
                kirjoittaja,
            )

        # print("Keskimäärin grafeemeja virkkeissä:")

        for kirjoittaja in korpus:
            # Alustetaan muuttujat keskiarvon laskua varten
            virkkeetKirjoittaja = 0
            merkkeja_kirjoittaja = 0

            for teksti in korpus[kirjoittaja]:
                # Grafeemeja ovat myös välimerkit, joten niitä ei tarvitse poistaa
                merkkeja_tekstissa = len(korpus[kirjoittaja][teksti])
                merkkeja_kirjoittaja += merkkeja_tekstissa

                virkkeitaTekstissa = len(
                    tokenisointi.tokenisoi_lauseet(korpus[kirjoittaja][teksti])
                )
                virkkeetKirjoittaja += virkkeitaTekstissa
                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    round(merkkeja_tekstissa / virkkeitaTekstissa, 2),
                    "Grafeemeja virkkeissä",
                    teksti,
                )
            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                round(merkkeja_kirjoittaja / virkkeetKirjoittaja, 2),
                "Grafeemeja virkkeissä",
                kirjoittaja,
            )

    # Funktio erikoismerkkien määrän laskemiseen teksteittäin
    def laske_erikoismerkit(korpus):
        for kirjoittaja in korpus:
            for teksti in korpus[kirjoittaja]:
                merkit = 0
                for merkki in korpus[kirjoittaja][teksti]:
                    if merkki.isalpha() == False and merkki != " " and merkki != "\n":
                        merkit += 1

    # Tulostetaan x määrä yleisintä sanaa
    def laske_hfl(korpus, maara):
        hfl = {}
        for kirjoittaja in korpus:
            hfl[kirjoittaja] = {}
            for teksti in korpus[kirjoittaja]:
                rivi = 2
                tokenit = tokenisointi.tokenisoi_sanat(korpus[kirjoittaja][teksti])
                dist = nltk.FreqDist(sana.lower() for sana in tokenit)
                yleisimmat = dist.most_common(maara)

                # Muutetaan tuple merkkijonoksi
                for tup in yleisimmat:
                    arvo = " : "
                    arvo = arvo.join(map(str, tup))

                    Openpyxl.lisaa_hfl_sivulle(arvo, rivi, teksti)
                    rivi += 1

    # Yule K'n funktio kertoo kielen rikkaudesta. Mitä suurempi luku, sitä vähemmän
    # kompleksisuutta. Yule I:ssä käänteinen: korkeampi luku, enemmän rikkautta.
    def laske_yule(teksti, otsikko, valinta):
        # M1: kaikkien saneiden määrä
        # M2: summa kaikkien lemmojen tuloista, (lemma * esiintymät^2)
        esiintymat = {}
        tokenit = tokenisointi.tokenisoi_sanat(teksti)

        for sane in tokenit:
            try:
                lemma = uralicApi.lemmatize(sane, "fin", word_boundaries=False)
                if len(lemma) > 0:
                    try:
                        avain = lemma[0]
                        esiintymat[avain] += 1
                    except KeyError:
                        esiintymat[avain] = 1
            except:
                print("Lemmaus epäonnistui, sane: ", sane)

        M1 = float(len(esiintymat))
        M2 = sum(
            [
                len(list(g)) * (freq**2)
                for freq, g in groupby(sorted(esiintymat.values()))
            ]
        )

        if valinta == "k":
            try:
                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin", 10000 * (M2 - M1) / (M1 * M1), "Yule K", otsikko
                )
            except ZeroDivisionError:
                print("Virhe Yule K:n laskennassa.")
        else:
            try:
                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin", (M1 * M1) / (M2 - M1), "Yule I", otsikko
                )

            except ZeroDivisionError:
                print("Virhe Yule I:n laskennassa.")

    # Käydään korpus teksteittäin läpi ja kutsutaan apufunktiota laskuihin
    # https://swizec.com/blog/measuring-vocabulary-richness-with-python/
    def alusta_yule(korpus, kirjain):
        for kirjoittaja in korpus:
            for teksti in korpus[kirjoittaja]:
                if kirjain == "K":
                    Laskut.laske_yule(korpus[kirjoittaja][teksti], teksti, "k")
                else:
                    Laskut.laske_yule(korpus[kirjoittaja][teksti], teksti, "i")

    # Funktio, jossa mistä tahansa listasta A voidaan poistaa lista B
    def poista_lista(tokenit, poistettavat):
        lista = [sana for sana in tokenit if sana not in poistettavat]
        return lista

    # Apufunktio kosinisimilaarisuuden laskentaan
    def kirjailijapipeline(nlp, teksti):
        erisnimet = uralic.erisnimetTeksti(teksti, nlp)
        tokenit = tokenisointi.tokenisoi_sanat(teksti)
        ilman_erisnimia = Laskut.poista_lista(tokenit, erisnimet)
        lemmat = []
        for sane in ilman_erisnimia:
            try:
                lemma = uralicApi.lemmatize(sane, "fin", word_boundaries=False)
                if len(lemma) > 0:
                    lemmat.append(lemma[0])
            except:
                print("Lemmaus epäonnistui, sane: ", sane)
        return lemmat

    # Kosinisimilaarisuudella lasketaan kahden tekstin samankaltaisuutta.
    # Vertailtavista teksteistä on erisnimet poistettu ja sanat lemmattu.
    def laske_kosinisimilaarisuus(korpus, nlp):
        data = []
        nimet = []

        # Yhdistetään kaikki saman kirjailijan tekstit samaan muuttujaan
        for kirjoittaja in korpus:
            nimet.append(kirjoittaja)
            tekstit = ""
            for teksti in korpus[kirjoittaja]:
                tekstit = tekstit + korpus[kirjoittaja][teksti]
            sanalista = Laskut.kirjailijapipeline(nlp, tekstit)
            data.append(" ".join(sanalista))

        Tfidf_vektori = TfidfVectorizer()
        vektorimatriisi = Tfidf_vektori.fit_transform(data)

        kosinimatriisi = cosine_similarity(vektorimatriisi)
        df = pd.DataFrame(data=kosinimatriisi, index=nimet, columns=nimet)
        Openpyxl.df_taulukkoon("Kosinisimilaarisuus", df)
        # Tulostetaan per teksti ja kirjoittaja löydettyjen sanaluokkien määrät

    def laske_sanaluokat(korpus):
        for kirjoittaja in korpus:
            saneita_kirjoittaja = 0

            # Sanaluokkamuuttujat (UD2) per kirjoittaja
            adj = 0
            adv = 0
            noun = 0
            verb = 0
            num = 0
            part = 0
            pron = 0
            acr = 0

            for teksti in korpus[kirjoittaja]:
                # Sanaluokkamuuttujat tekstille
                adjTeksti = 0
                advTeksti = 0
                nounTeksti = 0
                verbTeksti = 0
                numTeksti = 0
                partTeksti = 0
                pronTeksti = 0
                acrTeksti = 0

                saneet = tokenisoi_sanat(korpus[kirjoittaja][teksti])

                # Lasketaan saneiden määrä teksteittäin...
                saneita_tekstissa = len(saneet)

                # ...ja samalla kasvatetaan koko kirjoittajan saneiden määrää.
                saneita_kirjoittaja += saneita_tekstissa

                for sane in saneet:
                    try:
                        # Morfologinen analyysi uralicNLP:n avulla,
                        # josta erotellaan sanaluokka
                        tulos = uralicApi.analyze(sane, "fin")
                        if len(tulos) > 0:
                            analyysi = tulos[0][0]
                            osat = analyysi.split("+", 2)
                            lauseenjasen = ""
                            if len(osat) > 1:
                                lauseenjasen = osat[1]
                                match lauseenjasen:
                                    case "A":
                                        adj += 1
                                        adjTeksti += 1
                                    case "ACR":
                                        acr += 1
                                        acrTeksti += 1
                                    case "Adv":
                                        adv += 1
                                        advTeksti += 1
                                    case "N":
                                        noun += 1
                                        nounTeksti += 1
                                    case "Num":
                                        num += 1
                                        numTeksti += 1
                                    case "V":
                                        verb += 1
                                        verbTeksti += 1
                                    case "Pcle":
                                        part += 1
                                        partTeksti += 1

                                    # En tiedä mikä on Pcle#voi, mutta niitä löytyi paljon :D
                                    case "Pcle#voi":
                                        part += 1
                                        partTeksti += 1
                                    case "Po":
                                        adj += 1
                                        adjTeksti += 1
                                    case "Pron":
                                        pron += 1
                                        pronTeksti += 1
                    except:
                        print(
                            "Virhe sanaluokkajäsennyksessä. Vian aiheuttanut sane: ",
                            sane,
                        )

                print(kolmas_taso, "Adj ", adjTeksti)
                print(
                    kolmas_taso,
                    "Adj (%) ",
                    round((adjTeksti) / saneita_tekstissa * 100, 2),
                    " %",
                )
                print(kolmas_taso, "ACR ", acrTeksti)
                print(
                    kolmas_taso,
                    "ACR (%) ",
                    round(acrTeksti / saneita_tekstissa * 100, 2),
                    " %",
                )
                print(kolmas_taso, "N ", nounTeksti)
                print(
                    kolmas_taso,
                    "N (%) ",
                    round(nounTeksti / saneita_tekstissa * 100, 2),
                    " %",
                )
                print(kolmas_taso, "Num ", numTeksti)
                print(
                    kolmas_taso,
                    "Num (%) ",
                    round(numTeksti / saneita_tekstissa * 100, 2),
                    " %",
                )
                print(kolmas_taso, "V ", verbTeksti)
                print(
                    kolmas_taso,
                    "V (%) ",
                    round(verbTeksti / saneita_tekstissa * 100, 2),
                    " %",
                )
                print(kolmas_taso, "Pcle & Pcle#voi", partTeksti)
                print(
                    kolmas_taso,
                    "Pcle & Pcle#voi (%) ",
                    round(partTeksti / saneita_tekstissa * 100, 2),
                    " %",
                )
                print(kolmas_taso, "Pron ", pronTeksti)
                print(
                    kolmas_taso,
                    "Pron (%) ",
                    round(pronTeksti / saneita_tekstissa * 100, 2),
                    " %",
                )

        print(toka_taso, "Blogissa: ")
        print(kolmas_taso, "Adj ", round((adj) / saneita_kirjoittaja * 100, 2), " %")
        print(kolmas_taso, "ACR ", round(acr / saneita_kirjoittaja * 100, 2), " %")
        print(kolmas_taso, "N ", round(noun / saneita_kirjoittaja * 100, 2), " %")
        print(kolmas_taso, "Num ", round(num / saneita_kirjoittaja * 100, 2), " %")
        print(kolmas_taso, "V ", round(verb / saneita_kirjoittaja * 100, 2), " %")
        print(kolmas_taso, "Pcle & Pcle#voi", round(part / saneita_kirjoittaja * 100, 2), " %")
        print(kolmas_taso, "Pron ", round(pron / saneita_kirjoittaja * 100, 2), " %")

    # Etsitään erisnimet tekstistä.
    # Helmi on manuaalisesti myös luonut erisnimitiedoston, mutta
    # tämä funktio pätee kaikkiin manhdollisiin teksitiedostoihin.
    def laske_erisnimet_teksti(teksti, nlp):
        tagit = nlp(teksti)

        # Lista kaikista NER-tägeistä
        ners = [
            "PERSON",
            "GPE",
            "LOC",
            "DATE",
            "TIME",
            "MONEY",
            "QUANTITY",
            "ORDINAL" "CARDINAL",
        ]

        erisnimet = []
        for sana in tagit.ents:
            if sana.label_ in ners:
                erisnimet.append(sana.text)

        return erisnimet

    # Käydään korpus läpi silmukassa ja muodostetaan jokaiselle
    # tekstille lista siinä esiintyvistä erisnimistä
    def laske_erisnimet_korpus(korpus):
        nlp = spacy.load("fi_core_news_sm")
        tulos = {}
        for kirjoittaja in korpus:

            print(eka_taso, kirjoittaja)
            tulos[kirjoittaja] = {}
            for teksti in korpus[kirjoittaja]:
                print(toka_taso, teksti)

                erisnimet = Laskut.laske_erisnimet_teksti(teksti, nlp)

                print(kolmas_taso, erisnimet)
                korpus[kirjoittaja][teksti] = erisnimet

        return tulos

    #
    def laske_hapaxlegomenon(korpus):
        for kirjoittaja in korpus:
            # Käydään tekstit yksi kerrallaan läpi.
            for teksti in korpus[kirjoittaja]:
                saneet = tokenisoi_sanat(korpus[kirjoittaja][teksti])

                # Alustetaan muuttuja, johon kerätään yhden tekstin kaikki lemmat.
                lemmat = []

                # Käydään tekstin saneet kerrallaan läpi.
                for sane in saneet:
                    try:
                        # Lemmataan jokainen sane.
                        lemma = uralicApi.lemmatize(
                            sane.lower(), "fin", word_boundaries=False
                        )
                        if len(lemma) > 0:
                            # Lisätään sane lemmalistaan.
                            lemmat.append(lemma[0])

                        # Mikäli lemmaa ei löydy, lisätään sane sellaisenaan tyyppilistaan.
                        elif len(lemma) == 0:
                            lemmat.append(sane)
                    except:
                        print(
                            "Virhe lemmauksessa. Vian aiheuttanut sane: ",
                            sane,
                        )

                # Haetaan uniikit lemmat saneista muuttamalla lemmalista setiksi.
                tyypit = set(lemmat)
                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin", len(tyypit), "Hapax legomenon", teksti
                )

    # TTR = tekstin tyypit (= uniikit sanat) / tokeneilla (saneilla)
    # Mitä lähempänä TTR on numeroa 1, sitä enemmän sanastollista monipuolisuutta.
    def laske_ttr(korpus):
        for kirjoittaja in korpus:
            # Koko kirjoittajaa koskevat laskurit.
            saneita_kirjoittaja = 0
            tyypit_kirjoittaja = 0

            # Käydään tekstit yksi kerrallaan läpi.
            for teksti in korpus[kirjoittaja]:
                saneet = tokenisoi_sanat(korpus[kirjoittaja][teksti])

                # Lasketaan saneiden määrä teksteittäin...
                saneita_teksti = len(saneet)

                # ...ja samalla kasvatetaan koko kirjoittajan saneiden määrää.
                saneita_kirjoittaja += saneita_teksti
                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin", saneita_teksti, "Saneita", teksti
                ),

                # Alustetaan muuttuja, johon kerätään yhden tekstin kaikki lemmat.
                lemmat = []

                # Käydään tekstin saneet kerrallaan läpi.
                for sane in saneet:
                    try:
                        # Lemmataan jokainen sane.
                        lemma = uralicApi.lemmatize(
                            sane.lower(), "fin", word_boundaries=False
                        )
                        if len(lemma) > 0:
                            # Lisätään sane lemmalistaan.
                            lemmat.append(lemma[0])

                        # Mikäli lemmaa ei löydy, lisätään sane sellaisenaan tyyppilistaan.
                        elif len(lemma) == 0:
                            lemmat.append(sane)
                    except:
                        print(
                            kolmas_taso,
                            "Virhe lemmauksessa. Vian aiheuttanut sane: ",
                            sane,
                        )

                # Haetaan uniikit lemmat saneista muuttamalla lemmalista setiksi.
                tyypit = set(lemmat)
                tyyppeja = len(tyypit)
                Openpyxl.lisaa_taulukkoon("Teksteittäin", tyyppeja, "Tyyppejä", teksti),
                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin", tyyppeja / saneita_teksti, "TTR", teksti
                ),

                tyypit_kirjoittaja += len(tyypit)

            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin", saneita_teksti, "Saneita", kirjoittaja
            ),
            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin", tyypit_kirjoittaja, "Tyyppejä", kirjoittaja
            ),
            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                tyypit_kirjoittaja / saneita_kirjoittaja,
                "TTR",
                kirjoittaja,
            ),

    # Sanoissa morfeemien määrien keskiarvo, moodi, mediaani, minimi, maksimi,
    # vaihteluväli, keskihajonta ja varianssi per teksti ja kirjoittaja
    def laske_morfeemit(korpus):
        for kirjoittaja in korpus:
            morf_kirjoittaja = 0
            sanoja_kirjoittaja = 0

            # Listamuuttuja, johon laitetaan kaikkien kirjoittajan tekstien sanojen morfeemien määrät.
            morf_kirjoittaja = []

            for teksti in korpus[kirjoittaja]:
                sanoja_teksti = len(tokenisoi_sanat(korpus[kirjoittaja][teksti]))
                sanoja_kirjoittaja += sanoja_teksti
                morf_teksti = 0

                # Listamuuttuja, johon laitetaan kaikki tekstin sanojen morfeemien määrät.
                morf_teksti_lista = []

                # Tokenisoidaan teksti sanoiksi.
                tokenit = tokenisoi_sanat(korpus[kirjoittaja][teksti])

                # Käydään tokenit kerralla läpi ja uralicAPI-kirjaston avulla etsitään niistä morfeemit.

                for token in tokenit:
                    try:
                        morfeemit = uralicApi.segment(token, "fin")
                        if len(morfeemit) > 0:
                            # Funktio palauttaa kaikki sanan mahdolliset morfeemiyhdistelmät.
                            # Laskennan ja analyysin helpottamiseksi valitaan aina ensimmäinen vaihtoehto.
                            morf_teksti += len(morfeemit[0])
                            morf_kirjoittaja += len(morfeemit[0])

                            # Lisätään löydettyjen morfeemien määrät listoihin.
                            morf_kirjoittaja.append(len(morfeemit[0]))
                            morf_teksti_lista.append(len(morfeemit[0]))
                    except:
                        print(
                            kolmas_taso,
                            "Virhe morfologisessa analyysissa. Vian aiheuttanut tokeni: ",
                            token,
                        )

                # Tekstikohtaiset tilastot.
                minMorf = min(morf_teksti_lista)
                maxMorf = max(morf_teksti_lista)

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    round(statistics.mean(morf_teksti_lista), 2),
                    "Morfeemeja (avg)",
                    teksti,
                )

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    statistics.median(morf_teksti_lista),
                    "Morfeemeja (mediaani)",
                    teksti,
                )

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    statistics.mode(morf_teksti_lista),
                    "Morfeemeja (moodi)",
                    teksti,
                )

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin", minMorf, "Morfeemeja (min)", teksti
                )

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin", maxMorf, "Morfeemeja (max)", teksti
                )

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    maxMorf - minMorf,
                    "Morfeemeja (vaihteluväli)",
                    teksti,
                )

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    statistics.stdev(morf_teksti_lista),
                    "Morfeemeja (keskihajonta)",
                    teksti,
                )

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    statistics.variance(morf_teksti_lista),
                    "Morfeemeja (varianssi)",
                    teksti,
                )

            # Kirjoittajaikohtaiset tilastot.
            min_morf_kirjoittaja = min(morf_kirjoittaja)
            max_morf_kirjoittaja = max(morf_kirjoittaja)
            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                statistics.mean(morf_kirjoittaja),
                "Morfeemeja (avg)",
                kirjoittaja,
            )

            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                statistics.median(morf_kirjoittaja),
                "Morfeemeja (mediaani)",
                kirjoittaja,
            )

            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                statistics.mode(morf_kirjoittaja),
                "Morfeemeja (moodi)",
                kirjoittaja,
            )

            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin", min_morf_kirjoittaja, "Morfeemeja (min)", kirjoittaja
            )

            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin", max_morf_kirjoittaja, "Morfeemeja (max)", kirjoittaja
            )

            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                max_morf_kirjoittaja - min_morf_kirjoittaja,
                "Morfeemeja (vaihteluväli)",
                kirjoittaja,
            )

            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                statistics.stdev(morf_kirjoittaja),
                "Morfeemeja (keskihajonta)",
                kirjoittaja,
            )

            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                statistics.variance(morf_kirjoittaja),
                "Morfeemeja (varianssi)",
                kirjoittaja,
            )

    # Lemmausfunktio. jossa jokaisen tekstin sana lemmataan.
    # Mikäli sanalle tarjoaan useampaa lemmaa,
    # valitaan ensimmäinen. Virheitä siis tulee, mutta ohjelma
    # porskuttaa eteenmpäin.
    def lemmaa_korpus(korpus):
        lemmat_korpus = {}
        for kirjoittaja in korpus:
            lemmat_korpus[kirjoittaja] = {}
            for teksti in korpus[kirjoittaja]:
                # Tokenisoidaan teksti sanalistaksi
                saneet = tokenisoi_sanat(korpus[kirjoittaja][teksti])

                # Lista, johon tallennetaan yhden tekstin saneet
                lemmat_teksti = []
                for sane in saneet:
                    try:
                        lemma = uralicApi.lemmatize(sane, "fin", word_boundaries=False)
                        if len(lemma) > 0:
                            lemmat_teksti.append(lemma[0])
                    except:
                        print("Lemmaus epäonnistui, sane: ", sane)

                lemmat_korpus[kirjoittaja][teksti] = lemmat_teksti

        return lemmat_korpus

    # Sanaston tiheys
    def laske_tiheys(korpus):
        tiedosto = open(
            os.path.join(os.path.dirname(__file__), "sanalistat/funktiosanat.txt"), "r"
        )
        data = tiedosto.read()
        # Muutetaan tiedosto listaksi
        funktiosanat = data.split("\n")

        # Lemmataan korpus, jotta funktiosanoja voidaan vertailla ja poistaa
        lemmat = Laskut.lemmaa_korpus(korpus)

        for kirjoittaja in lemmat:
            for teksti in lemmat[kirjoittaja]:
                maara = 0
                for lemma in lemmat[kirjoittaja][teksti]:
                    if lemma in funktiosanat:
                        maara += 1

                Openpyxl.lisaa_taulukkoon("Teksteittäin", maara, "Funktiosanoja", teksti)
                sanamaara = len(lemmat[kirjoittaja][teksti])
                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    round(maara / sanamaara * 100, 2),
                    "Funktiosanoja (%)",
                    teksti,
                )

    def tilastoja(asetukset, korpus):
        nlp = spacy.load("fi_core_news_sm")

        if asetukset["Sanamäärä"] == True:
            Laskut.sanamaara(korpus)
        if asetukset["Virkemäärä"] == True:
            Laskut.virkemaara(korpus)
        if asetukset["Grafeemeja sanoissa ja virkkeissä"] == True:
            Laskut.laske_grafeemeina(korpus)
        if asetukset["Virkepituus sanoina"] == True:
            Laskut.laske_sanat_virkkeissa(korpus)
        # if asetukset["erikoismerkit"] == True:
        #     erikoismerkit(korpus)
        if asetukset["HFL 50"] == True:
            Laskut.laske_hfl(korpus, 50)
        if asetukset["HFL 100"] == True:
            Laskut.laske_hfl(korpus, 100)
        if asetukset["Yule K"] == True:
            Laskut.alusta_yule(korpus, "K")
        if asetukset["Yule I"] == True:
            Laskut.alusta_yule(korpus, "I")
        if asetukset["Kosinisimilaarisuus"] == True:
            Laskut.laske_kosinisimilaarisuus(korpus, nlp)
        if asetukset["Morfeemeja sanoissa"] == True:
            Laskut.laske_morfeemit(korpus)
        if asetukset["Type-token ratio"] == True:
            Laskut.laske_ttr(korpus)
        if asetukset["Sanaluokkien frekvenssit"] == True:
            Laskut.laske_sanaluokat(korpus)
        if asetukset["Hapax Legomena"] == True:
            Laskut.laske_hapaxlegomenon(korpus)
        if asetukset["Erisnimet"] == True:
            Laskut.laske_erisnimet_korpus(korpus)
        if asetukset["Sanaston tiheys"] == True:
            Laskut.laske_tiheys(korpus)

if __name__ == "__main__":
    sovellus = Sovellus()
    sovellus.mainloop()
