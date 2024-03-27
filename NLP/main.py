import openpyxl as xl
import laskut
import uralic
import os
import hae
from openpyxl.utils.dataframe import dataframe_to_rows

# ----------- LOGIIKKA  -----------
# Tässä tiedostossa kutsutaan funktioita ja kirjoitetaan kaikki
# Excel-tiedostoon

# ----------- EXCEL-TIEDOSTON LUONTI -----------
# Tarkistetaan, onko tiedosto jo olemassa.
polku = "./tulokset.xlsx"
if os.path.isfile(polku) == True:
    workbook = xl.load_workbook("tulokset.xlsx")
else:
    workbook = xl.Workbook()
    workbook = xl.Workbook()
    teksteittain = workbook["Sheet"]
    teksteittain.title = "Teksteittäin"
    workbook.create_sheet("HFL")
    workbook.create_sheet("Kosinisimilaarisuus")    


# ----------- GLOBAALIT  MUUTTUJAT -----------
rivisanakirja = {}
sarakesanakirja = {}

# Tulostiedoston rivien otsikot
rivinimet = ["Sanamäärä", "Virkemäärä", "Grafeemeja sanoissa", "Grafeemeja virkkeissä", 
             "Morfeemeja (avg)", 
             "Morfeemeja (mediaani)", 
             "Morfeemeja (moodi)", 
             "Morfeemeja (min)", 
             "Morfeemeja (max)", 
             "Morfeemeja (vaihteluväli)", 
             "Morfeemeja (keskihajonta)", 
             "Morfeemeja (varianssi)", 
             "Virkepituus sanoina (avg)",
             "Virkepituus sanoina (keskihajonta)",
             "Saneita",
             "Tyyppejä",
             "TTR", 
             "Sanaston tiheys",
             "Funktiosanoja",
             "Funktiosanoja (%)" ,
             "Hapax legomenon", "Yule I", "Yule K"]

# ----------- WORKBOOK-FUNKTIOT -----------


def tekstinimet(korpus):
    nimet = []
    for kirjoittaja in korpus:
        for teksti in korpus[kirjoittaja]:
            nimet.append(teksti)
        nimet.append(kirjoittaja)
    return nimet

def luoSarakenimet(korpus):
    sarakkeet = {}
    nykyinen = 2
    for kirjoittaja in korpus:
        for teksti in korpus[kirjoittaja]:
            sarakkeet[teksti] = nykyinen
            nykyinen += 1
        sarakkeet[kirjoittaja] = nykyinen
        nykyinen += 1
        
    return sarakkeet

def luoRivinimet():
    rivit = {}
    nykyinen = 2

    for rivinimi in rivinimet:
        rivit[rivinimi] = nykyinen
        nykyinen += 1
    return rivit

# Luodaan taulukon perusmuoto
def alustaTaulukko(korpus):
    workbook.active = workbook["Teksteittäin"]
    sivu = workbook.active

    # Alustetaan rivien (eli asetusten) nimet
    rivinro = 2
    for rivi in rivinimet:
        solu = sivu.cell(row=rivinro, column=1)
        solu.value = rivi
        rivinro += 1

    # Alustetaan sarakkeiden (eli tekstien ja blogien) nimet
    otsikot = tekstinimet(korpus)
    sarakenro = 2
    for sarake in otsikot:
        solu = sivu.cell(row=1, column=sarakenro)
        solu.value = sarake
        sarakenro += 1

def alustaHFL(korpus):
    workbook.active = workbook["HFL"]
    sivu = workbook.active
    # Alustetaan sarakkeiden (eli tekstien ja blogien) nimet
    otsikot = tekstinimet(korpus)
    sarakenro = 1
    for sarake in otsikot:
        solu = sivu.cell(row=1, column=sarakenro)
        solu.value = sarake
        sarakenro += 1

def lisaaTaulukkoon(sivunimi, arvo, rivi, sarake):
    global workbook
    workbook.active = workbook[sivunimi]
    sivu = workbook.active
    solu = sivu.cell(row=rivisanakirja[rivi], column=sarakesanakirja[sarake])
    solu.value = arvo

def lisaaHFL(arvo, rivi, sarake):
    global workbook
    workbook.active = workbook["HFL"]
    sivu = workbook.active
    solu = sivu.cell(row=rivi, column=sarakesanakirja[sarake])
    
    solu.value = arvo

def dataframeTaulukkoon(sivunimi, data):
    global workbook
    workbook.active = workbook[sivunimi]
    sivu = workbook.active
    for rivi in dataframe_to_rows(data, index=True, header=True):
        sivu.append(rivi)

# ----------- WORKBOOK-FUNKTIOT -----------

def logiikka(asetukset):

    # Noudetaan korpus
    polku = os.path.abspath(os.path.curdir)
    kansio = "corpora"
    korpus = hae.hae_korpus(os.path.join(polku, kansio))

    global rivisanakirja
    global sarakesanakirja

    # Alustetaan sarakkeiden/rivien nimet sanakirjaan, jotta tietoja voidaan 
    # tallentaa tiettyyn sarakkeeseen (tekstiin/blogiin) tai riviin.
    # Solun arvon saa komennolla row_cells[sarakenimet["Jokin"]].value
    sarakesanakirja = luoSarakenimet(korpus)
    rivisanakirja = luoRivinimet()
    alustaTaulukko(korpus)
    alustaHFL(korpus)

    # Aloitetaan laskut
    kutsut = [laskut.tilastoja(asetukset, korpus), uralic.uralicAsetukset(asetukset, korpus)]

    for kutsu in kutsut:
        kutsu

    workbook.save("tulokset.xlsx")
    print("Valmista")