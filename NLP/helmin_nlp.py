import hae
import nltk
import openpyxl as xl
import os
import pandas as pd
import spacy
import statistics
import sys
import time
import tokenisointi
import tkinter as tk
import ttkbootstrap as ttk

from functools import partial
from itertools import groupby
from openpyxl.utils.dataframe import dataframe_to_rows
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from threading import Thread
from tokenisointi import tokenisoi_sanat
from ttkbootstrap.dialogs import Messagebox
from ttkbootstrap.constants import *
from uralicNLP import uralicApi
from time import sleep
from threading import Thread
from queue import Queue
from enum import Enum, auto


class Korpus:
    korpus = {}
    asetukset = {}

    # Haetaan korpus ja lasketaan sen koko
    def alusta_korpus():
        """
        Haetaan korpus sanakirjaksi ja lasketaan korpuksen koko
        """
        # Noudetaan korpus
        polku = os.path.abspath(os.path.curdir)
        kansio = "corpora"
        Korpus.korpus = hae.hae_korpus(os.path.join(polku, kansio))

    def valitse(arvo, asetus):
        """
        Asetusten päivittämisfunktio
        """
        global asetukset
        Korpus.asetukset[asetus] = arvo.get()

    def hae_kasiteltavat_yht(dct, yht=0):
        """
        Funktio, jossa lasketaan kaikkien käsiteltävien entiteettien määrä,
        eli kansiot ja niissä olevat tiedostot rekursiivisesti
        """
        yht = yht + len(dct)
        if isinstance(dct, dict):
            for avain1, arvo1 in dct.items():
                if isinstance(arvo1, dict):
                    yht = Korpus.hae_kasiteltavat_yht(arvo1, yht)
        return yht

    def hae_laskut_yht():
        """
        Funktio, jossa lasketaan asetusten perusteella, montako laskua/vaihetta
        ohjelmassa on. Esimerkiksi käyttäjä on valinnut kohdat Sanamäärä,
        Virkemäärä ja kosinisimilaarisuus jolloin laskuja on hae_kasiteltavat_yht() x 2 +
        len(korpus), koska viimeisessä laskuissa käsitellään koko kirjoittajaa,
        ei yksittäisiä tekstejä.
        """
        yht = 0
        kasiteltavia = Korpus.hae_kasiteltavat_yht(Korpus.korpus)
        kirjoittajia = len(Korpus.korpus)
        for asetus in Korpus.asetukset:
            if Korpus.asetukset[asetus] == True:
                if (
                    asetus == "Kosinisimilaarisuus"
                ):
                    yht += kirjoittajia
                elif (
                    asetus == "Sanaston tiheys"
                    or asetus == "Hapax Legomena"
                    or asetus == "Yule K"
                    or asetus == "Yule I"
                    or asetus == "HFL 50"
                    or asetus == "HFL 100"
                ):
                    yht += kasiteltavia - kirjoittajia
                else:
                    yht += kasiteltavia
        return yht


class LippuTarkoitus(Enum):
    PAIVITA_EDISTYMIS_TEKSTI = auto()
    PAIVITA_EDISTYMIS_LUKU = auto()
    TIEDOSTO_AUKI_ALOITUS = auto()
    TIEDOSTO_AUKI_LOPETUS = auto()


class Lippu:
    """
    Lipulla päivitetään Tkinteriä toisesta säikeestä, eli laskusäikeestä.
    Lippu voi päivittää joko latauspalkin tekstiä tai arvoa.
    """

    def __init__(self, lippu_tyyppi: LippuTarkoitus, lippu_arvo: str):
        self.lippu_tyyppi = lippu_tyyppi
        self.lippu_arvo = lippu_arvo


class Paaikkuna(ttk.Window):  # Super class
    def __init__(self):
        super().__init__(themename="minty")
        self.withdraw()
        self.splash = Splash(self)

        # Alustetaan ikkuna
        self.title("Helmin NLP-ohjelma")
        self.iconbitmap("99_85283.ico")

        # Alustetaan ikkunan koko suhteessa näytön kokoon
        leveys = self.winfo_screenwidth()
        korkeus = self.winfo_screenheight()
        self.position_center()
        self.geometry("%sx%s" % (int(leveys / 1.5), int(korkeus / 1.5)))
        self.minsize(int(leveys / 1.5), int(korkeus / 1.5))

        # Haetaan korpus
        Korpus.alusta_korpus()

        # Odotetaan hetki, jotta kaikki varmasti on valmista
        time.sleep(4)

        self.splash.destroy()
        self.deiconify()

        # ------------ PÄÄIKKUNAN OSIOT ------------
        # Ikkuna jakautuu kolmeen osioon: ylhällä valikkopalkkiin,
        # vasemmalla asetuksiin, ja oikealla infolaatikkoon
        valikkopalkki = Valikkopalkki(self)
        self.config(menu=valikkopalkki)

        self.asetus_frame = Asetusikkuna(self)
        self.ohje_frame = Ohjeikkuna(self)

    def vaihda_vari(vari):
        """
        Funktio sovelluksen värin muuttamiseen.
        """
        tyyli = ttk.Style()
        tyyli.theme_use(vari)

    def create_frame_buttons(self) -> ttk.Frame:
        """
        Create and return a frame that contains buttons
        """

        self.frame_buttons = ttk.Frame(self)

        self.btn_download = ttk.Button(
            self.frame_buttons, text="Download", command=self.on_download_button_clicked
        )

        self.btn_test = ttk.Button(
            self.frame_buttons, text="Test", command=self.on_test_button_clicked
        )

        self.lbl_status = ttk.Label(self.frame_buttons)

        self.btn_download.pack()
        self.btn_test.pack()
        self.lbl_status.pack()

        return self.frame_buttons

    def on_test_button_clicked(self):
        print("Test")

    def on_download_button_clicked(self):
        new_thread = Thread(
            target=self.suorita,
            args=("sky.jpg",),
            daemon=True,  # Kun pääohjelma suljetaan, myös tämä säie sulkeutuu
        )
        new_thread.start()

    def suorita(self, file_name: str):
        """
        Ajetaan ohjelma erillisessä säikeessä
        """
        for progress in range(1, 101):  # goes up to 100 (%)
            Lippu = Lippu(
                lippu_tyyppi=LippuTarkoitus.PAIVITA_EDISTYMIS_TEKSTI,
                lippu_arvo=f"Downloading {file_name}...{progress} %",
            )
            self.jono_viesti.put(Lippu)
            self.event_generate(
                "<<TarkistaJono>>"
            )  # virtual event with an arbitary name
            sleep(1)

        Lippu = Lippu(
            lippu_tyyppi=LippuTarkoitus.PAIVITA_EDISTYMIS_TEKSTI,
            lippu_arvo="Finished downloading!",
        )
        self.jono_viesti.put(Lippu)
        self.event_generate("<<TarkistaJono>>")

        # ----------- FUNKTIOT -----------

        # Kaikkien osioiden valintalaatikot luodaan
        # listasta silmukassa.


class Asetusikkuna(ttk.Frame):
    """
    Vasemmanpuoleinen osio, josta löytyy kolmen ala-otsikon alta
    kaikki käyttäjälle mahdolliset ohjelman asetukset.
    """

    def __init__(self, parent):
        super().__init__(parent)

        # MUUTTUJAT
        self.prog_yht = 0
        self.vaiheita = 0

        # ULKOASU

        self.grid(row=0, column=0, sticky="ns", pady=20)
        ttk.Label(self, text="Asetukset", font=("Garamond", 28)).pack()

        vaihtoehdot_sanakirja = {
            "Perustilastoja": [
                "Sanamäärä",
                "Virkemäärä",
                "Grafeemeja sanoissa ja virkkeissä",
                "Morfeemeja sanoissa",
                "Virkepituus sanoina",
            ],
            "Sanasto": [
                "Type-token ratio",
                "Sanaston tiheys",
                "HFL 50",
                "HFL 100",
                "Hapax Legomena",
            ],
            "Kirjoittajien vertailu": ["Yule K", "Yule I", "Kosinisimilaarisuus"],
        }

        for osio in vaihtoehdot_sanakirja:
            Asetusosio(self, osio, vaihtoehdot_sanakirja[osio])

        self.kaynnistys_nappi = ttk.Button(
            self, text="Käynnistä", command=self.kaynnista
        )
        self.kaynnistys_nappi.pack()

        ttk.Style().configure(
            "custom.Horizontal.TProgressbar",
            thickness=15,
            bordercolor="black",
            borderwidth="1",
            pbarrelief="flat",
        )

        self.edistymispalkki = ttk.Progressbar(
            self,
            mode="determinate",
            orient="horizontal",
            length=200,
            style="custom.Horizontal.TProgressbar",
        )

        self.edistymispalkki.pack_forget()

        self.jono_viesti = Queue()

        self.ladattu = ttk.Label(self)
        self.ladattu.pack_forget()

        self.kasiteltava = ttk.Label(self)
        self.kasiteltava.pack_forget()

        self.bind("<<CheckQueue>>", self.tarkista_jono)

    def kaynnista(self):
        """
        Näytetään edistymispalkki ja aloitetaan laskut omassa säikeessään.
        """
        tarkistus = Asetusikkuna.tarkista_onko_auki()
        if tarkistus == True:
            Asetusikkuna.luo_varoitus(self, "Aloitus")
            print("Sulje Excel-tiedosto")
        else:
            self.nayta_edistymispalkki
            saie = Thread(target=self.aja_ohjelma, daemon=True)
            saie.start()

    def nayta_edistymispalkki(self, vaiheita):
        """
        Ennen ohjelman käynnistymistä edistymispalkki on piilotettu.
        Ohjelman käynnistyttyä palkki näytetään, maksimiarvona
        käyttäjän asetuksiin perustuen kaikkien laskuen määrä.
        """
        self.edistymispalkki.config(maximum=vaiheita, value=0)
        self.edistymispalkki.pack(pady=10)

        self.ladattu.pack()
        self.kasiteltava.pack()

    def tarkista_onko_auki():
        try:
            tiedosto = open("tulokset.xlsx", "r")
            return False

        except:
            return True


    def aja_ohjelma(self):
        """
        Ajetaan ohjelma erillisessä säikeessä, jotta GUI ei jäädy.
        """
        self.kaynnistys_nappi.config(state="disabled")
        self.vaiheita = Korpus.hae_laskut_yht()
        self.prog_yht = 0
        self.nayta_edistymispalkki(self.vaiheita)
        Openpyxl.luo_workbook()

        aloitus_lippu = Lippu(
                lippu_tyyppi=LippuTarkoitus.PAIVITA_EDISTYMIS_LUKU,
                lippu_arvo=f"Valmiina 0 / {self.vaiheita}",
        )
        self.jono_viesti.put(aloitus_lippu)
        self.event_generate("<<CheckQueue>>")

        Laskut.tilastoja(Korpus.asetukset, self)

        self.merkitse_aloitetuksi("")

        self.kaynnistys_nappi.config(state="normal")

        self.lopeta_ohjelma()

    def lopeta_ohjelma(self):
        print("Lopeta ohjelma")
        tarkistus = Asetusikkuna.tarkista_onko_auki()
        if tarkistus == True:
            print("Lopetus: Tiedosto auki")
            tarkistus_lippu = Lippu(
                    lippu_tyyppi=LippuTarkoitus.TIEDOSTO_AUKI_LOPETUS,
                    lippu_arvo=""
            )
            self.jono_viesti.put(tarkistus_lippu)
            self.event_generate("<<CheckQueue>>")
        else:
            print("Lopetus: tiedosto kiinni")
            Openpyxl.tallenna_tulokset()       

    def merkitse_aloitetuksi(self, viesti):
        lippu = Lippu(LippuTarkoitus.PAIVITA_EDISTYMIS_TEKSTI, lippu_arvo=viesti)
        self.jono_viesti.put(lippu)
        self.event_generate("<<CheckQueue>>")

    def merkitse_tehdyksi(self):
        self.prog_yht += 1
        lippu = Lippu(
            lippu_tyyppi=LippuTarkoitus.PAIVITA_EDISTYMIS_LUKU,
            lippu_arvo=f"Valmiina {self.prog_yht} / {self.vaiheita}",
        )
        self.jono_viesti.put(lippu)
        self.event_generate("<<CheckQueue>>")

    def tarkista_jono(self, event):  # Pääsäikeessä
        """
        Luetaan jono, jonka avulla päivitetään käyttöliittymän
        latauspalkkia.
        """
        viesti: Lippu  # Type hint
        viesti = self.jono_viesti.get()

        if viesti.lippu_tyyppi == LippuTarkoitus.PAIVITA_EDISTYMIS_LUKU:
            self.ladattu.configure(text=viesti.lippu_arvo)
            self.edistymispalkki.config(value=self.prog_yht)
        if viesti.lippu_tyyppi == LippuTarkoitus.PAIVITA_EDISTYMIS_TEKSTI:
            self.kasiteltava.configure(text=viesti.lippu_arvo)
        if viesti.lippu_tyyppi == LippuTarkoitus.TIEDOSTO_AUKI_LOPETUS:
            Asetusikkuna.luo_varoitus(self, "Lopetus")


    def luo_varoitus(self, vaihe):
        nappi_arvo = Messagebox.retrycancel(message="Sulje tulokset.xlsx-tiedosto jaksaaksesi",
                                title="Virhe!",
                                alert=True)

        if nappi_arvo == "Retry":
            if vaihe == "Lopetus":
                Asetusikkuna.lopeta_ohjelma(self)
            else:
                Asetusikkuna.kaynnista(self)
        

class Asetusosio(ttk.Frame):
    """
    Luokka, jossa luodaan jokaiselle asetusten alaluokalle otsikko
    ja valintalaatikot sanakirjasta.
    """

    def __init__(self, parent, osio_nimi, vaihtoehdot):
        super().__init__(master=parent)

        # Gridin ulkoasu
        self.rowconfigure(0, weight=1)
        self.columnconfigure((0, 1), weight=1, uniform="a")

        # Widgetit
        ttk.Label(self, text=osio_nimi, font=("Garamond 15 italic")).grid(
            row=0, column=0, sticky="nsew"
        )
        self.luo_valintaruudut(vaihtoehdot).grid(row=0, column=1, sticky="nsew")

        # Sijoitetaan AsetusOsio
        self.pack(expand=True, fill="both", padx=30, pady=30)

    def luo_valintaruudut(self, vaihtoehdot):
        """
        Funktio kaikkien valintaruutujen luomiseen.
        """
        global asetukset
        frame = ttk.Frame(master=self)
        for vaihtoehto in vaihtoehdot:
            Korpus.asetukset[vaihtoehto] = False
            checkbox_var = tk.BooleanVar()
            checkbox_var.set(False)
            ttk.Checkbutton(
                frame,
                variable=checkbox_var,
                text=vaihtoehto,
                command=partial(Korpus.valitse, checkbox_var, vaihtoehto),
            ).pack(expand=True, fill="both")
        return frame


class Ohjeikkuna(ttk.Frame):
    """
    Pääikkunan oikea puoli, jossa ohjelman käyttöohjeet.
    """

    def __init__(self, parent):
        super().__init__(parent)
        self.grid(row=0, column=1, sticky="ns", pady=20)

        ttk.Label(self, text="Ohjeet", font=("Garamond", 28)).grid(row=0)

        ohje = """Helmin NLP-ohjelma on työkalu suomenkielisten tekstien yksinkertaiseen analyysiin ja vertailuun. 
                      
Ohjelman suorituskansiosta löytyy "corpora"-niminen kansio, johon voit tiputtaa haluamasi UTF-8 -muotoiset tekstit. Mikäli haluat vertailla useampaa kirjoittajaa, luo jokaiselle kirjailijalle oma alikansio. Älä kuitenkaan luo kirjailijan alakansion sisälle sisempiä alakansioita. Kansioiden ja tiedostojen nimillä ei ole merkitystä. 
                      
Kaikki tilastot ja vertailujen tulokset tallennetaan samaan, myöskin ohjelman suorituskansiosta löytyvään tiedostoon nimeltä "tulokset.xlsx."

Tulevia toimintoja:
    - lauseiden tunnistaminen ja analyysi
    - painike ohjelman suorituksen keskeyttämiseksi
    - yleistä optimointia
    - sanaluokkien tallentaminen tiedostoon               
                      """

        teksti = ttk.Text(
            self,
            wrap=WORD,
        )
        teksti.grid(row=1)
        teksti.insert(END, ohje)
        teksti.config(state=DISABLED)  # Vain-luku -tilainen teksti


class Valikkopalkki(ttk.Menu):
    """
    Yläreunan valikkopalkki. Käyttäjä voi lukea ohjelmassa käytetyistä
    kirjastoista ja ohjelman tekjöistä, sekä vaihtaa ohjelman väriteemaa.
    """

    def __init__(self, parent):
        super().__init__(parent)

        # ----------- INFO ----------
        info_vaihtoehdot = ttk.Menu(self, tearoff=False)
        self.add_cascade(label="Info", menu=info_vaihtoehdot)
        info_vaihtoehdot.add_command(
            label="Tietoa sovelluksesta", command=self.nayta_info
        )
        info_vaihtoehdot.add_command(label="Sulje", command=self.lopeta)

        # ----------- SOVELLUKSEN VÄRIN VAIHTAMINEN ----------
        vari_menu = ttk.Menu(self, tearoff=False)
        self.add_cascade(label="Teema", menu=vari_menu)
        for x in ["flatly", "minty", "morph", "solar", "superhero", "darkly"]:
            vari_menu.add_command(label=x, command=lambda x=x: Paaikkuna.vaihda_vari(x))

    def nayta_info(self):
        Messagebox.ok(
            """Sovelluksessa käytetyt kirjastot:
NLTK, Openypxl, Pandas, Scikit-learn, Spacy, TTKBootstrap, UralicNLP.
Github: https://github.com/ellaelomaa/nlp
Tekijät: Ella Elomaa ja Helmi Siiroinen (2024)
"""
        )

    def lopeta(self):
        sys.exit(0)


class Splash(ttk.Toplevel):
    """
    Splash-ikkuna, joka näytetään ennen varsinaista ohjelmaa.
    Splash-ikkunan ollessa näkyvissä tervehditään käyttäjää ja taustalla ladataan korpus valmiiksi.
    """

    def __init__(self, parent):
        ttk.Toplevel.__init__(self, parent)
        self.title("Splash")

        self.position_center()
        self.geometry("200x200")

        teksti = "Tervetuloa! Pieni hetki, sovellus käynnistyy."
        splash_label = ttk.Text(self, wrap=WORD)
        splash_label.insert(INSERT, teksti)
        splash_label.pack()

        self.update()


class Laskut:
    """
    Säie, jossa tehdään kaikki teksteihin liittyvät laskut. Käytetään toista
    säiettä, jotta käyttöliittymä ei jäädy laskujen ajaksi.

    Jokaisen tekstin/kirjoittajan käsittelyn loputtua ilmoitetaan tehtävä tehdyksi GUI:lle.
    """

    def laske_sanamaarat(self):
        """
        Kokonaissanamäärät per teksti ja kirjoittaja
        """
        self.merkitse_aloitetuksi("Lasketaan sanamääriä")
        for kirjoittaja in Korpus.korpus:
            sanojaKirjoittaja = 0
            for teksti in Korpus.korpus[kirjoittaja]:
                sanojaTekstissa = len(
                    tokenisointi.tokenisoi_sanat(Korpus.korpus[kirjoittaja][teksti])
                )
                sanojaKirjoittaja += sanojaTekstissa

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin", sanojaTekstissa, "Sanamäärä", teksti
                )
                self.merkitse_tehdyksi()
            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin", sanojaKirjoittaja, "Sanamäärä", kirjoittaja
            )
            self.merkitse_tehdyksi()

    def laske_virkemaarat(self):
        """
        Virkkeiden määrä per teksti ja kirjoittaja
        """
        self.merkitse_aloitetuksi("Lasketaan virkemääriä")
        for kirjoittaja in Korpus.korpus:
            lauseita_kirjoittaja = 0
            for teksti in Korpus.korpus[kirjoittaja]:
                lauseitaTekstissa = len(
                    tokenisointi.tokenisoi_virkkeet(Korpus.korpus[kirjoittaja][teksti])
                )
                lauseita_kirjoittaja += lauseitaTekstissa

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin", lauseitaTekstissa, "Virkemäärä", teksti
                )
                self.merkitse_tehdyksi()
            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin", lauseita_kirjoittaja, "Virkemäärä", kirjoittaja
            )
            self.merkitse_tehdyksi()

    def laske_sanat_virkkeissa(self):
        """
        Sanojen määrien keskiarvo ja keskihajonta per teksti ja kirjoittaja
        """
        self.merkitse_aloitetuksi("Lasketaan sanoja virkkeissä")
        for kirjoittaja in Korpus.korpus:
            sanoja_kirjoittaja = []
            for teksti in Korpus.korpus[kirjoittaja]:
                virkkeet = tokenisointi.tokenisoi_virkkeet(
                    Korpus.korpus[kirjoittaja][teksti]
                )
                sanoja_virkkeissa = []
                for virke in virkkeet:
                    virke = tokenisointi.tokenisoi_sanat(virke)
                    sanoja_virkkeissa.append(len(virke))
                    sanoja_kirjoittaja.append(len(virke))

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    round(statistics.mean(sanoja_virkkeissa), 2),
                    "Virkepituus sanoina (avg)",
                    teksti,
                )
                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    round(statistics.stdev(sanoja_virkkeissa), 2),
                    "Virkepituus sanoina (keskihajonta)",
                    teksti,
                )
                self.merkitse_tehdyksi()
            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                round(statistics.mean(sanoja_kirjoittaja), 2),
                "Virkepituus sanoina (avg)",
                kirjoittaja,
            )
            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                round(statistics.stdev(sanoja_kirjoittaja), 2),
                "Virkepituus sanoina (keskihajonta)",
                kirjoittaja,
            )
            self.merkitse_tehdyksi()

    def laske_grafeemeina(self):
        """
        Grafeemien (=merkkien) määrä per teksti ja kirjoittaja
        """
        # Käydään korpus kirjoittaja kerrallaan läpi
        self.merkitse_aloitetuksi("Lasketaan grafeemeja")

        for kirjoittaja in Korpus.korpus:
            # Alustetaan muuttujat kirjoittajaa koskeville muuttujille.
            sanoja_kirjoittaja = 0
            merkkeja_kirjoittaja = 0

            # Käydään kirjoittajan tekstit kerrallaan läpi
            for teksti in Korpus.korpus[kirjoittaja]:
                merkkeja_tekstissa = 0

                # Keskiarvoa varten lasketaan sanojen määrä tekstissä
                sanalista = tokenisointi.tokenisoi_sanat(
                    Korpus.korpus[kirjoittaja][teksti]
                )
                sanoja_tekstissa = len(sanalista)

                # Kasvatetaan samalla koko kirjoittajan sanamäärää laskevaa muuttujaa
                sanoja_kirjoittaja += len(sanalista)

                # Lasketaan, kuinka pitkiä sanat ovat teksteittäin
                for sana in sanalista:
                    merkkeja_tekstissa += len(sana)
                    merkkeja_kirjoittaja += len(sana)

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    round(merkkeja_tekstissa / sanoja_tekstissa, 2),
                    "Grafeemeja sanoissa",
                    teksti,
                )
                self.merkitse_tehdyksi()
            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                round(merkkeja_kirjoittaja / sanoja_kirjoittaja, 2),
                "Grafeemeja sanoissa",
                kirjoittaja,
            )
            self.merkitse_tehdyksi()

        for kirjoittaja in Korpus.korpus:
            # Alustetaan muuttujat keskiarvon laskua varten
            virkkeetKirjoittaja = 0
            merkkeja_kirjoittaja = 0

            for teksti in Korpus.korpus[kirjoittaja]:
                # Grafeemeja ovat myös välimerkit, joten niitä ei tarvitse poistaa
                merkkeja_tekstissa = len(Korpus.korpus[kirjoittaja][teksti])
                merkkeja_kirjoittaja += merkkeja_tekstissa

                virkkeitaTekstissa = len(
                    tokenisointi.tokenisoi_lauseet(Korpus.korpus[kirjoittaja][teksti])
                )
                virkkeetKirjoittaja += virkkeitaTekstissa
                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    round(merkkeja_tekstissa / virkkeitaTekstissa, 2),
                    "Grafeemeja virkkeissä",
                    teksti,
                )
            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                round(merkkeja_kirjoittaja / virkkeetKirjoittaja, 2),
                "Grafeemeja virkkeissä",
                kirjoittaja,
            )

    def laske_erikoismerkit():
        """
        Funktio erikoismerkkien määrän laskemiseen teksteittäin
        """
        for kirjoittaja in Korpus.korpus:
            for teksti in Korpus.korpus[kirjoittaja]:
                merkit = 0
                for merkki in Korpus.korpus[kirjoittaja][teksti]:
                    if merkki.isalpha() == False and merkki != " " and merkki != "\n":
                        merkit += 1

    def laske_hfl(maara, self):
        # TODO: Poista erisnimet
        # TODO: normalisointi
        # TODO: Poista funktiosanat
        """
        Lasketaan 50 tai sata yleisintä sanaa tekstissä
        """
        self.merkitse_aloitetuksi("Lasketaan yleisimpiä sanoja")
        hfl = {}
        for kirjoittaja in Korpus.korpus:
            hfl[kirjoittaja] = {}
            for teksti in Korpus.korpus[kirjoittaja]:
                rivi = 2
                tokenit = tokenisointi.tokenisoi_sanat(
                    Korpus.korpus[kirjoittaja][teksti]
                )
                dist = nltk.FreqDist(sana.lower() for sana in tokenit)
                yleisimmat = dist.most_common(maara)

                # Muutetaan tuple merkkijonoksi
                for tup in yleisimmat:
                    arvo = " : "
                    arvo = arvo.join(map(str, tup))
                    rivi += 1

                    Openpyxl.lisaa_hfl_sivulle(arvo, rivi, teksti)
                self.merkitse_tehdyksi()

    def laske_yule(teksti, otsikko, valinta, self):
        """
        Yule K'n funktio kertoo kielen rikkaudesta. Mitä suurempi luku, sitä vähemmän
        kompleksisuutta. Yule I:ssä käänteinen: korkeampi luku, enemmän rikkautta.
        """
        # M1: kaikkien saneiden määrä
        # M2: summa kaikkien lemmojen tuloista, (lemma * esiintymät^2)
        esiintymat = {}
        tokenit = tokenisointi.tokenisoi_sanat(teksti)
        self.merkitse_aloitetuksi("Lasketaan Yule")

        for sane in tokenit:
            try:
                lemma = uralicApi.lemmatize(sane, "fin", word_boundaries=False)
                if len(lemma) > 0:
                    try:
                        avain = lemma[0]
                        esiintymat[avain] += 1
                    except KeyError:
                        esiintymat[avain] = 1
            except:
                print("Lemmaus epäonnistui, sane: ", sane)

        M1 = float(len(esiintymat))
        M2 = sum(
            [
                len(list(g)) * (freq**2)
                for freq, g in groupby(sorted(esiintymat.values()))
            ]
        )

        if valinta == "k":
            try:
                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin", 10000 * (M2 - M1) / (M1 * M1), "Yule K", otsikko
                )
                self.merkitse_tehdyksi()
            except ZeroDivisionError:
                print("Virhe Yule K:n laskennassa.")
        else:
            try:
                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin", (M1 * M1) / (M2 - M1), "Yule I", otsikko
                )
                self.merkitse_tehdyksi()

            except ZeroDivisionError:
                print("Virhe Yule I:n laskennassa.")

    # Käydään korpus teksteittäin läpi ja kutsutaan apufunktiota laskuihin
    # https://swizec.com/blog/measuring-vocabulary-richness-with-python/
    def alusta_yule(kirjain, self):
        for kirjoittaja in Korpus.korpus:
            for teksti in Korpus.korpus[kirjoittaja]:
                if kirjain == "K":
                    Laskut.laske_yule(
                        Korpus.korpus[kirjoittaja][teksti], teksti, "k", self
                    )
                else:
                    Laskut.laske_yule(
                        Korpus.korpus[kirjoittaja][teksti], teksti, "i", self
                    )

    def poista_lista(tokenit, poistettavat):
        """
        Funktio, jossa mistä tahansa listasta A voidaan poistaa lista B
        """
        lista = [sana for sana in tokenit if sana not in poistettavat]
        return lista

    def kosini_pipeline(nlp, teksti):
        """
        Apufunktio kosinisimilaarisuuden laskentaan
        """
        erisnimet = Laskut.hae_erisnimet_teksti(teksti, nlp)
        tokenit = tokenisointi.tokenisoi_sanat(teksti)
        ilman_erisnimia = Laskut.poista_lista(tokenit, erisnimet)
        lemmat = []
        for sane in ilman_erisnimia:
            try:
                lemma = uralicApi.lemmatize(sane, "fin", word_boundaries=False)
                if len(lemma) > 0:
                    lemmat.append(lemma[0])
            except:
                print("Lemmaus epäonnistui, sane: ", sane)
        return lemmat

    def laske_kosinisimilaarisuus(nlp, self):
        """
        Kosinisimilaarisuudella lasketaan kahden tekstin samankaltaisuutta.
        Vertailtavista teksteistä on erisnimet poistettu ja sanat lemmattu.
        """
        data = []
        nimet = []
        self.merkitse_aloitetuksi("Lasketaan kosinisimilaarisuutta")

        # Yhdistetään kaikki saman kirjailijan tekstit samaan muuttujaan
        for kirjoittaja in Korpus.korpus:
            nimet.append(kirjoittaja)
            tekstit = ""
            for teksti in Korpus.korpus[kirjoittaja]:
                tekstit = tekstit + Korpus.korpus[kirjoittaja][teksti]
            sanalista = Laskut.kosini_pipeline(nlp, tekstit)
            data.append(" ".join(sanalista))
            self.merkitse_tehdyksi()

        Tfidf_vektori = TfidfVectorizer()
        vektorimatriisi = Tfidf_vektori.fit_transform(data)

        kosinimatriisi = cosine_similarity(vektorimatriisi)
        df = pd.DataFrame(data=kosinimatriisi, index=nimet, columns=nimet)
        Openpyxl.df_taulukkoon("Kosinisimilaarisuus", df)
        # Tulostetaan per teksti ja kirjoittaja löydettyjen sanaluokkien määrät

    def laske_sanaluokat(korpus, self):
        """
        Lasketaan yleisimpien sanaluokkien esiintymät ja prosenttimäärä
        per teksti ja kirjoittaja
        """
        self.merkitse_aloitetuksi("Lasketaan sanaluokkia")
        for kirjoittaja in korpus:
            saneita_kirjoittaja = 0

            # Sanaluokkamuuttujat (UD2) per kirjoittaja
            adj = 0
            adv = 0
            noun = 0
            verb = 0
            num = 0
            part = 0
            pron = 0
            acr = 0

            for teksti in korpus[kirjoittaja]:
                # Sanaluokkamuuttujat tekstille
                adjTeksti = 0
                advTeksti = 0
                nounTeksti = 0
                verbTeksti = 0
                numTeksti = 0
                partTeksti = 0
                pronTeksti = 0
                acrTeksti = 0

                saneet = tokenisoi_sanat(korpus[kirjoittaja][teksti])

                # Lasketaan saneiden määrä teksteittäin...
                saneita_tekstissa = len(saneet)

                # ...ja samalla kasvatetaan koko kirjoittajan saneiden määrää.
                saneita_kirjoittaja += saneita_tekstissa

                for sane in saneet:
                    try:
                        # Morfologinen analyysi uralicNLP:n avulla,
                        # josta erotellaan sanaluokka
                        tulos = uralicApi.analyze(sane, "fin")
                        if len(tulos) > 0:
                            analyysi = tulos[0][0]
                            osat = analyysi.split("+", 2)
                            lauseenjasen = ""
                            if len(osat) > 1:
                                lauseenjasen = osat[1]
                                match lauseenjasen:
                                    case "A":
                                        adj += 1
                                        adjTeksti += 1
                                    case "ACR":
                                        acr += 1
                                        acrTeksti += 1
                                    case "Adv":
                                        adv += 1
                                        advTeksti += 1
                                    case "N":
                                        noun += 1
                                        nounTeksti += 1
                                    case "Num":
                                        num += 1
                                        numTeksti += 1
                                    case "V":
                                        verb += 1
                                        verbTeksti += 1
                                    case "Pcle":
                                        part += 1
                                        partTeksti += 1

                                    # En tiedä mikä on Pcle#voi, mutta niitä löytyi paljon :D
                                    case "Pcle#voi":
                                        part += 1
                                        partTeksti += 1
                                    case "Po":
                                        adj += 1
                                        adjTeksti += 1
                                    case "Pron":
                                        pron += 1
                                        pronTeksti += 1
                    except:
                        print(
                            "Virhe sanaluokkajäsennyksessä. Vian aiheuttanut sane: ",
                            sane,
                        )

                print("Adj ", adjTeksti)
                print(
                    "Adj (%) ",
                    round((adjTeksti) / saneita_tekstissa * 100, 2),
                    " %",
                )
                print("ACR ", acrTeksti)
                print(
                    "ACR (%) ",
                    round(acrTeksti / saneita_tekstissa * 100, 2),
                    " %",
                )
                print("N ", nounTeksti)
                print(
                    "N (%) ",
                    round(nounTeksti / saneita_tekstissa * 100, 2),
                    " %",
                )
                print("Num ", numTeksti)
                print(
                    "Num (%) ",
                    round(numTeksti / saneita_tekstissa * 100, 2),
                    " %",
                )
                print("V ", verbTeksti)
                print(
                    "V (%) ",
                    round(verbTeksti / saneita_tekstissa * 100, 2),
                    " %",
                )
                print("Pcle & Pcle#voi", partTeksti)
                print(
                    "Pcle & Pcle#voi (%) ",
                    round(partTeksti / saneita_tekstissa * 100, 2),
                    " %",
                )
                print("Pron ", pronTeksti)
                print(
                    "Pron (%) ",
                    round(pronTeksti / saneita_tekstissa * 100, 2),
                    " %",
                )

        print("Blogissa: ")
        print("Adj ", round((adj) / saneita_kirjoittaja * 100, 2), " %")
        print("ACR ", round(acr / saneita_kirjoittaja * 100, 2), " %")
        print("N ", round(noun / saneita_kirjoittaja * 100, 2), " %")
        print("Num ", round(num / saneita_kirjoittaja * 100, 2), " %")
        print("V ", round(verb / saneita_kirjoittaja * 100, 2), " %")
        print("Pcle & Pcle#voi", round(part / saneita_kirjoittaja * 100, 2), " %")
        print("Pron ", round(pron / saneita_kirjoittaja * 100, 2), " %")

    def hae_erisnimet_teksti(teksti, nlp):
        """
        Funktio erisnimien määrän laskemiseen tekstissä. Palauttaa listan.
        """
        tagit = nlp(teksti)

        # Lista kaikista NER-tägeistä
        ners = [
            "PERSON",
            "GPE",
            "LOC",
            "DATE",
            "TIME",
            "MONEY",
            "QUANTITY",
            "ORDINAL" "CARDINAL",
        ]

        erisnimet = []
        for sana in tagit.ents:
            if sana.label_ in ners:
                erisnimet.append(sana.text)

        return erisnimet

    def laske_erisnimet_korpus():
        """
        Käydään korpus läpi silmukassa ja muodostetaan jokaiselle tekstille
        lista siinä esiintyvistä erisnimistä. Palauttaa sanakirjan.
        """
        nlp = spacy.load("fi_core_news_sm")
        tulos = {}
        for kirjoittaja in Korpus.korpus:
            tulos[kirjoittaja] = {}
            for teksti in Korpus.korpus[kirjoittaja]:
                erisnimet = Laskut.hae_erisnimet_teksti(teksti, nlp)
                Korpus.korpus[kirjoittaja][teksti] = erisnimet

        return tulos

    def laske_hapaxlegomenon(self):
        """
        Lasketaan jokaiselle tekstille HFL, eli uniikkien lemmojen määrä.
        """
        self.merkitse_aloitetuksi("Lasketaan hapax legomena")
        for kirjoittaja in Korpus.korpus:
            # Käydään tekstit yksi kerrallaan läpi.
            for teksti in Korpus.korpus[kirjoittaja]:
                saneet = tokenisoi_sanat(Korpus.korpus[kirjoittaja][teksti])

                # Alustetaan muuttuja, johon kerätään yhden tekstin kaikki lemmat.
                lemmat = []

                # Käydään tekstin saneet kerrallaan läpi.
                for sane in saneet:
                    try:
                        # Lemmataan jokainen sane.
                        lemma = uralicApi.lemmatize(
                            sane.lower(), "fin", word_boundaries=False
                        )
                        if len(lemma) > 0:
                            # Lisätään sane lemmalistaan.
                            lemmat.append(lemma[0])

                        # Mikäli lemmaa ei löydy, lisätään sane sellaisenaan tyyppilistaan.
                        elif len(lemma) == 0:
                            lemmat.append(sane)
                    except:
                        print(
                            "Virhe lemmauksessa. Vian aiheuttanut sane: ",
                            sane,
                        )

                # Haetaan uniikit lemmat saneista muuttamalla lemmalista setiksi.
                tyypit = set(lemmat)
                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin", len(tyypit), "Hapax legomenon", teksti
                )
                self.merkitse_tehdyksi()

    def laske_ttr(self):
        """
        Laskee type-token ration (TTR) per teksti ja kirjoittaja. Mitä lähempänä
        TTR on lukua 1, sitä enemmän sanastollista monipuolisuutta.
        """
        self.merkitse_aloitetuksi("Lasketaan TTR")
        for kirjoittaja in Korpus.korpus:
            # Koko kirjoittajaa koskevat laskurit.
            saneita_kirjoittaja = 0
            tyypit_kirjoittaja = 0

            # Käydään tekstit yksi kerrallaan läpi.
            for teksti in Korpus.korpus[kirjoittaja]:
                saneet = tokenisoi_sanat(Korpus.korpus[kirjoittaja][teksti])

                # Lasketaan saneiden määrä teksteittäin...
                saneita_teksti = len(saneet)

                # ...ja samalla kasvatetaan koko kirjoittajan saneiden määrää.
                saneita_kirjoittaja += saneita_teksti
                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin", saneita_teksti, "Saneita", teksti
                )

                # Alustetaan muuttuja, johon kerätään yhden tekstin kaikki lemmat.
                lemmat = []

                # Käydään tekstin saneet kerrallaan läpi.
                for sane in saneet:
                    try:
                        # Lemmataan jokainen sane.
                        lemma = uralicApi.lemmatize(
                            sane.lower(), "fin", word_boundaries=False
                        )
                        if len(lemma) > 0:
                            # Lisätään sane lemmalistaan.
                            lemmat.append(lemma[0])

                        # Mikäli lemmaa ei löydy, lisätään sane sellaisenaan tyyppilistaan.
                        elif len(lemma) == 0:
                            lemmat.append(sane)
                    except:
                        print(
                            "Virhe lemmauksessa. Vian aiheuttanut sane: ",
                            sane,
                        )

                # Haetaan uniikit lemmat saneista muuttamalla lemmalista setiksi.
                tyypit = set(lemmat)
                tyyppeja = len(tyypit)
                tyypit_kirjoittaja += len(tyypit)

                Openpyxl.lisaa_taulukkoon("Teksteittäin", tyyppeja, "Tyyppejä", teksti)
                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin", tyyppeja / saneita_teksti, "TTR", teksti
                )

                self.merkitse_tehdyksi()

            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin", saneita_teksti, "Saneita", kirjoittaja
            )
            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin", tyypit_kirjoittaja, "Tyyppejä", kirjoittaja
            )
            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                tyypit_kirjoittaja / saneita_kirjoittaja,
                "TTR",
                kirjoittaja,
            )
            self.merkitse_tehdyksi()

    def laske_morfeemit(self):
        """
        Laskee morfeemien määrän keskiarvon, moodin, mediaanin, minimin,
        maksimin, vaihteluvälin, keskihajonnan ja varianssin per teksti
        ja kirjoittaja.
        """
        self.merkitse_aloitetuksi("Lasketaan morfeemeja")
        for kirjoittaja in Korpus.korpus:
            # Listamuuttuja, johon laitetaan kaikkien kirjoittajan tekstien sanojen morfeemien määrät.
            morf_kirjoittaja_lista = []

            for teksti in Korpus.korpus[kirjoittaja]:
                # Listamuuttuja, johon laitetaan kaikki tekstin sanojen morfeemien määrät.
                morf_teksti_lista = []

                # Tokenisoidaan teksti sanoiksi.
                tokenit = tokenisoi_sanat(Korpus.korpus[kirjoittaja][teksti])

                # Käydään tokenit kerralla läpi ja uralicAPI-kirjaston avulla etsitään niistä morfeemit.

                for token in tokenit:
                    try:
                        morfeemit = uralicApi.segment(token, "fin")

                        # Tarkistetaan, onnistuko sanan morfeemien löytäminen.
                        if len(morfeemit) > 0:
                            # Funktio palauttaa kaikki sanan mahdolliset morfeemiyhdistelmät.
                            # Laskennan ja analyysin helpottamiseksi valitaan aina ensimmäinen vaihtoehto.

                            # Lisätään löydettyjen morfeemien määrät listoihin.
                            morf_kirjoittaja_lista.append(len(morfeemit[0]))
                            morf_teksti_lista.append(len(morfeemit[0]))
                    except:
                        print(
                            "Virhe morfologisessa analyysissa. Vian aiheuttanut tokeni: ",
                            token,
                        )
                # Tekstikohtaiset tilastot.
                minMorf = min(morf_teksti_lista)
                maxMorf = max(morf_teksti_lista)

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    round(statistics.mean(morf_teksti_lista), 2),
                    "Morfeemeja (avg)",
                    teksti,
                )

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    statistics.median(morf_teksti_lista),
                    "Morfeemeja (mediaani)",
                    teksti,
                )

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    statistics.mode(morf_teksti_lista),
                    "Morfeemeja (moodi)",
                    teksti,
                )

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin", minMorf, "Morfeemeja (min)", teksti
                )

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin", maxMorf, "Morfeemeja (max)", teksti
                )

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    maxMorf - minMorf,
                    "Morfeemeja (vaihteluväli)",
                    teksti,
                )

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    statistics.stdev(morf_teksti_lista),
                    "Morfeemeja (keskihajonta)",
                    teksti,
                )

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    statistics.variance(morf_teksti_lista),
                    "Morfeemeja (varianssi)",
                    teksti,
                )
                self.merkitse_tehdyksi()

            # Kirjoittajaikohtaiset tilastot.
            min_morf_kirjoittaja = min(morf_kirjoittaja_lista)
            max_morf_kirjoittaja = max(morf_kirjoittaja_lista)
            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                statistics.mean(morf_kirjoittaja_lista),
                "Morfeemeja (avg)",
                kirjoittaja,
            )

            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                statistics.median(morf_kirjoittaja_lista),
                "Morfeemeja (mediaani)",
                kirjoittaja,
            )

            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                statistics.mode(morf_kirjoittaja_lista),
                "Morfeemeja (moodi)",
                kirjoittaja,
            )

            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin", min_morf_kirjoittaja, "Morfeemeja (min)", kirjoittaja
            )

            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin", max_morf_kirjoittaja, "Morfeemeja (max)", kirjoittaja
            )

            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                max_morf_kirjoittaja - min_morf_kirjoittaja,
                "Morfeemeja (vaihteluväli)",
                kirjoittaja,
            )

            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                statistics.stdev(morf_kirjoittaja_lista),
                "Morfeemeja (keskihajonta)",
                kirjoittaja,
            )

            Openpyxl.lisaa_taulukkoon(
                "Teksteittäin",
                statistics.variance(morf_kirjoittaja_lista),
                "Morfeemeja (varianssi)",
                kirjoittaja,
            )
            self.merkitse_tehdyksi()

    def lemmaa_korpus():
        """
        Funktio, jossa koko korpuksen saneet lemmataan. Mikäli sanalle tarjotaan
        useampaa lemmaa, palautetaan ensimmäinen. Virheitä siis tulee tapahtumaan.
        Palauttaa sanakirjan.
        """
        lemmat_korpus = {}
        for kirjoittaja in Korpus.korpus:
            lemmat_korpus[kirjoittaja] = {}
            for teksti in Korpus.korpus[kirjoittaja]:
                # Tokenisoidaan teksti sanalistaksi
                saneet = tokenisoi_sanat(Korpus.korpus[kirjoittaja][teksti])

                # Lista, johon tallennetaan yhden tekstin saneet
                lemmat_teksti = []
                for sane in saneet:
                    try:
                        lemma = uralicApi.lemmatize(sane, "fin", word_boundaries=False)
                        if len(lemma) > 0:
                            lemmat_teksti.append(lemma[0])
                    except:
                        print("Lemmaus epäonnistui, sane: ", sane)

                lemmat_korpus[kirjoittaja][teksti] = lemmat_teksti

        return lemmat_korpus

    def laske_tiheys(self):
        """
        Laskee funktiosanojen määrän ja sanaston tiheyden, eli kuinka suuri
        osa tekstistä on funktiosanoja.
        """
        self.merkitse_aloitetuksi("Lasketaan sanaston tiheyttä")
        tiedosto = open(
            os.path.join(os.path.dirname(__file__), "sanalistat/funktiosanat.txt"), "r"
        )
        data = tiedosto.read()
        # Muutetaan tiedosto listaksi
        funktiosanat = data.split("\n")

        # Lemmataan korpus, jotta funktiosanoja voidaan vertailla ja poistaa
        lemmat = Laskut.lemmaa_korpus()

        for kirjoittaja in lemmat:
            for teksti in lemmat[kirjoittaja]:
                maara = 0
                for lemma in lemmat[kirjoittaja][teksti]:
                    if lemma in funktiosanat:
                        maara += 1

                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin", maara, "Funktiosanoja", teksti
                )
                sanamaara = len(lemmat[kirjoittaja][teksti])
                Openpyxl.lisaa_taulukkoon(
                    "Teksteittäin",
                    round(maara / sanamaara * 100, 2),
                    "Funktiosanoja (%)",
                    teksti,
                )
                self.merkitse_tehdyksi()

    def tilastoja(asetukset, self):
        nlp = spacy.load("fi_core_news_sm")

        if asetukset["Sanamäärä"] == True:
            Laskut.laske_sanamaarat(self)
        if asetukset["Virkemäärä"] == True:
            Laskut.laske_virkemaarat(self)
        if asetukset["Grafeemeja sanoissa ja virkkeissä"] == True:
            Laskut.laske_grafeemeina(self)
        if asetukset["Virkepituus sanoina"] == True:
            Laskut.laske_sanat_virkkeissa(self)
        # if asetukset["erikoismerkit"] == True:
        #     erikoismerkit(korpus)
        if asetukset["HFL 50"] == True:
            Laskut.laske_hfl(50, self)
        if asetukset["HFL 100"] == True:
            Laskut.laske_hfl(100, self)
        if asetukset["Yule K"] == True:
            Laskut.alusta_yule("K", self)
        if asetukset["Yule I"] == True:
            Laskut.alusta_yule("I", self)
        if asetukset["Kosinisimilaarisuus"] == True:
            Laskut.laske_kosinisimilaarisuus(nlp, self)
        if asetukset["Morfeemeja sanoissa"] == True:
            Laskut.laske_morfeemit(self)
        if asetukset["Type-token ratio"] == True:
            Laskut.laske_ttr(self)
        # if asetukset["Sanaluokkien frekvenssit"] == True:
        #     Laskut.laske_sanaluokat()
        if asetukset["Hapax Legomena"] == True:
            Laskut.laske_hapaxlegomenon(self)
        # if asetukset["Erisnimet"] == True:
        #     Laskut.laske_erisnimet_korpus()
        if asetukset["Sanaston tiheys"] == True:
            Laskut.laske_tiheys(self)


class Openpyxl:
    """
    Luokka, jossa luodaan ja muokataan tulokset.xlsx-tiedostoa.
    """

    sarakesanakirja = {}
    rivisanakirja = {}
    workbook: xl.Workbook

    def luo_workbook():
        """
        Funktio, joka hakee olemassa olevan xlsx-tiedoston, tai sen puuttuessa
        luo sen.
        """
        # Tarkistetaan, onko tiedosto jo olemassa.
        polku = "./tulokset.xlsx"
        global workbook
        global rivisanakirja
        global sarakesanakirja

        rivisanakirja = Openpyxl.luo_rivinimet()
        sarakesanakirja = Openpyxl.luo_sarakenimet()

        if os.path.isfile(polku) == True:
            workbook = xl.load_workbook("tulokset.xlsx")
            workbook.save("tulokset.xlsx")
            return workbook
        else:  # Jos ei, niin luodaan tiedosto ja sen sivut
            # Korvataan ensimmäisen sivun oletusnimi
            workbook = xl.Workbook()
            teksteittain = workbook["Sheet"]
            teksteittain.title = "Teksteittäin"

            # Muut sivut
            workbook.create_sheet("HFL")
            workbook.create_sheet("Kosinisimilaarisuus")

            workbook = Openpyxl.alusta_taulukko()
            workbook = Openpyxl.alusta_hfl_sivu()
            return workbook

    # Sanakirjat rivien ja sarakkeiden nimille ja niiden indekseille.
    # OpenPyxl:ssä tiettyyn soluun voi lisätä rivin ja sarakkeen indeksin
    # mukaan, ei nimen, joten luodaan sanakirja muotoa nimi:indeksi sekä
    # riveille että sarakkeille.

    # Tulostiedoston rivien otsikot
    rivinimet = [
        "Sanamäärä",
        "Virkemäärä",
        "Grafeemeja sanoissa",
        "Grafeemeja virkkeissä",
        "Morfeemeja (avg)",
        "Morfeemeja (mediaani)",
        "Morfeemeja (moodi)",
        "Morfeemeja (min)",
        "Morfeemeja (max)",
        "Morfeemeja (vaihteluväli)",
        "Morfeemeja (keskihajonta)",
        "Morfeemeja (varianssi)",
        "Virkepituus sanoina (avg)",
        "Virkepituus sanoina (keskihajonta)",
        "Saneita",
        "Tyyppejä",
        "TTR",
        "Funktiosanoja",
        "Funktiosanoja (%)",
        "Hapax legomenon",
        "Yule I",
        "Yule K",
    ]

    # ----------- WORKBOOK-FUNKTIOT -----------

    def hae_tekstinimet():
        """
        Tulostaulukon sarakkeita varten haetaan corpus-kansiosta kaikki
        tekstien ja kirjoittajien nimet.
        """
        nimet = []
        for kirjoittaja in Korpus.korpus:
            for teksti in Korpus.korpus[kirjoittaja]:
                nimet.append(teksti)
            nimet.append(kirjoittaja)
        return nimet

    def luo_sarakenimet():
        """
        Sivuilla, joissa käsitellään vain kirjoittajia kokonaisuutena,
        sarakkeiden niminä on vain kirjoittajien (eli kansioiden) nimet.
        """
        sarakkeet = {}
        nykyinen = 2
        for kirjoittaja in Korpus.korpus:
            for teksti in Korpus.korpus[kirjoittaja]:
                sarakkeet[teksti] = nykyinen
                nykyinen += 1
            sarakkeet[kirjoittaja] = nykyinen
            nykyinen += 1

        return sarakkeet

    def luo_rivinimet():
        rivit = {}
        nykyinen = 2

        for rivinimi in Openpyxl.rivinimet:
            rivit[rivinimi] = nykyinen
            nykyinen += 1

        return rivit

    def alusta_taulukko():
        """
        Tulostaulukon ensimmäisen sivun luonti.
        """
        global workbook
        workbook.active = workbook["Teksteittäin"]
        sivu = workbook.active

        # Alustetaan rivien (eli asetusten) nimet
        rivinro = 2
        for rivi in Openpyxl.rivinimet:
            solu = sivu.cell(row=rivinro, column=1)
            solu.value = rivi
            rivinro += 1

        # Alustetaan sarakkeiden (eli tekstien ja kirjoittajien) nimet
        otsikot = Openpyxl.hae_tekstinimet()
        sarakenro = 2
        for sarake in otsikot:
            solu = sivu.cell(row=1, column=sarakenro)
            solu.value = sarake
            sarakenro += 1

        return workbook

    def alusta_hfl_sivu():
        """
        Taulukon toisen sivun (HFL) luonti.
        """
        global workbook
        workbook.active = workbook["HFL"]
        sivu = workbook.active

        # Alustetaan sarakkeiden (eli tekstien ja kirjoittajien) nimet
        otsikot = Openpyxl.hae_tekstinimet()
        sarakenro = 1
        for sarake in otsikot:
            solu = sivu.cell(row=1, column=sarakenro)
            solu.value = sarake
            sarakenro += 1

        return workbook

    def lisaa_taulukkoon(sivunimi, arvo, rivi, sarake):
        """
        Funktio, jolla lisätään ensimmäiselle sivulle dataa.
        """
        global workbook

        workbook.active = workbook[sivunimi]
        sivu = workbook.active
        solu = sivu.cell(row=rivisanakirja[rivi], column=sarakesanakirja[sarake])
        solu.value = arvo

    def lisaa_hfl_sivulle(arvo, rivi, sarake):
        """
        HFL-sivulla ei ole rivien otsikoita, ainoastaan eniten esiintyvät sanat
        ja niiden määrät.
        """
        global workbook
        workbook.active = workbook["HFL"]
        sivu = workbook.active
        solu = sivu.cell(row=rivi, column=sarakesanakirja[sarake] - 1)
        solu.value = arvo

    def df_taulukkoon(sivunimi, data):
        """
        Funktio, jossa pandasin Dataframe lisätään xlsx-tiedostoon.
        """
        global workbook
        workbook.active = workbook[sivunimi]
        sivu = workbook.active
        for rivi in dataframe_to_rows(data, index=True, header=True):
            sivu.append(rivi)

    def alusta_xlsx():
        """
        Alustetaan xlsx-tiedosto korpukseen perustuen.
        """
        Openpyxl.sarakesanakirja = Openpyxl.luo_sarakenimet()
        Openpyxl.rivisanakirja = Openpyxl.luo_rivinimet()
        Openpyxl.alusta_taulukko()
        Openpyxl.alusta_hfl_sivu()

    def tallenna_tulokset():
        global workbook
        workbook.save("tulokset.xlsx")


if __name__ == "__main__":
    main_window = Paaikkuna()
    main_window.mainloop()
